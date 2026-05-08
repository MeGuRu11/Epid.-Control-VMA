from __future__ import annotations

import json
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from app.application.services.exchange_service import EXCEL_SHEET_TITLES, ExchangeService
from app.application.services.form100_service_v2 import Form100ServiceV2
from app.infrastructure.db import models_sqlalchemy as models
from tests.integration.test_exchange_service_import_reports import make_session_factory, seed_actor
from tests.integration.test_form100_v2_service import make_create_request

FORM100_PDF_NOTE = "PDF-артефакты карточек Формы 100 экспортируются отдельно через Form100 ZIP"


def _seed_ismp_and_form100(session_factory, actor_id: int) -> str:
    with session_factory() as session:
        patient = models.Patient(full_name="Patient For Full Export", sex="M")
        session.add(patient)
        session.flush()
        emr_case = models.EmrCase(patient_id=patient.id, hospital_case_no="CASE-FULL-001")
        session.add(emr_case)
        session.flush()
        session.add(
            models.RefIsmpAbbreviation(
                code="ВАП",
                name="Вентилятор-ассоциированная пневмония",
                description="Тестовая справочная запись",
            )
        )
        session.add(
            models.IsmpCase(
                emr_case_id=emr_case.id,
                ismp_type="ВАП",
                start_date=date(2026, 5, 1),
            )
        )

    form100_service = Form100ServiceV2(session_factory=session_factory)
    created = form100_service.create_card(make_create_request(), actor_id=actor_id)
    return created.id


def test_full_json_includes_ismp_and_form100_with_nested_json(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "full_json_form100_ismp.db")
    actor_id = seed_actor(session_factory)
    form100_id = _seed_ismp_and_form100(session_factory, actor_id)
    service = ExchangeService(session_factory=session_factory)
    json_path = tmp_path / "full_export.json"

    result = service.export_json(json_path, exported_by="exchange_admin", actor_id=actor_id)

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    data = payload["data"]
    assert result["counts"]["ref_ismp_abbreviations"] == 1
    assert result["counts"]["ismp_case"] == 1
    assert result["counts"]["form100"] == 1
    assert result["counts"]["form100_data"] == 1
    assert payload["notes"]["form100_pdf"] == FORM100_PDF_NOTE
    assert data["ref_ismp_abbreviations"][0]["code"] == "ВАП"
    assert data["ismp_case"][0]["ismp_type"] == "ВАП"
    assert data["form100"][0]["id"] == form100_id

    form100_data = data["form100_data"][0]
    assert isinstance(form100_data["main_json"], dict)
    assert isinstance(form100_data["bodymap_annotations_json"], list)
    assert isinstance(form100_data["bodymap_tissue_types_json"], list)
    assert form100_data["main_json"]["main_full_name"] == "Ivan Ivanov"


def test_full_xlsx_includes_ismp_and_form100_sheets(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "full_xlsx_form100_ismp.db")
    actor_id = seed_actor(session_factory)
    _seed_ismp_and_form100(session_factory, actor_id)
    service = ExchangeService(session_factory=session_factory)
    xlsx_path = tmp_path / "full_export.xlsx"

    result = service.export_excel(xlsx_path, exported_by="exchange_admin", actor_id=actor_id)

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    for table_name in ("ref_ismp_abbreviations", "ismp_case", "form100", "form100_data"):
        assert result["counts"][table_name] == 1
        assert EXCEL_SHEET_TITLES[table_name] in workbook.sheetnames


def test_full_zip_manifest_notes_form100_pdf_exported_separately(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "full_zip_form100_note.db")
    actor_id = seed_actor(session_factory)
    _seed_ismp_and_form100(session_factory, actor_id)
    service = ExchangeService(session_factory=session_factory)
    zip_path = tmp_path / "full_export.zip"

    service.export_zip(zip_path, exported_by="exchange_admin", actor_id=actor_id)

    with zipfile.ZipFile(zip_path, "r") as zf:
        manifest = json.loads(zf.read("manifest.json"))

    assert manifest["notes"]["form100_pdf"] == FORM100_PDF_NOTE
