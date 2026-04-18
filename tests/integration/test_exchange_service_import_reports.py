from __future__ import annotations

import csv
import json
import zipfile
from collections.abc import Callable, Iterator
from contextlib import AbstractContextManager, contextmanager
from datetime import UTC, date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.application.services.exchange_service import ExchangeService
from app.infrastructure.db import models_sqlalchemy as models
from app.infrastructure.db.models_sqlalchemy import Base, User
from app.infrastructure.security.sha256 import sha256_file


def make_session_factory(db_path: Path) -> Callable[[], AbstractContextManager[Session]]:
    engine = create_engine(f"sqlite:///{db_path.as_posix()}", future=True)
    Base.metadata.create_all(engine)
    session_local = sessionmaker(
        bind=engine,
        autoflush=False,
        autocommit=False,
        expire_on_commit=False,
        future=True,
    )

    @contextmanager
    def _session_scope() -> Iterator[Session]:
        session: Session = session_local()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    return _session_scope


def seed_actor(session_factory: Callable[[], AbstractContextManager[Session]]) -> int:
    with session_factory() as session:
        actor = User(login="exchange_admin", password_hash="hash", role="admin", is_active=True)
        session.add(actor)
        session.flush()
        return int(actor.id)


def test_import_csv_returns_error_report_and_log(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "exchange_csv_report.db")
    actor_id = seed_actor(session_factory)
    service = ExchangeService(session_factory=session_factory)
    csv_path = tmp_path / "patients.csv"

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "full_name", "dob", "sex", "category", "military_unit", "military_district"])
        writer.writerow([1, "Иванов Иван", "2024-01-01", "M", "cat", "", ""])
        writer.writerow([2, "Петров Петр", "bad-date", "M", "cat", "", ""])

    result = service.import_csv(csv_path, "patients", actor_id=actor_id, mode="merge")

    assert result["count"] == 1
    assert result["error_count"] == 1
    assert result["summary"]["rows_total"] == 2
    assert result["summary"]["added"] == 1
    assert result["summary"]["errors"] == 1

    error_log_path = Path(str(result["error_log_path"]))
    assert error_log_path.exists()
    payload = json.loads(error_log_path.read_text(encoding="utf-8"))
    assert payload["errors_count"] == 1
    assert payload["errors"][0]["scope"] == "patients"

    with session_factory() as session:
        patients = session.query(models.Patient).all()
    assert len(patients) == 1
    assert str(patients[0].full_name) == "Иванов Иван"


def test_import_excel_returns_error_report_and_log(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "exchange_excel_report.db")
    actor_id = seed_actor(session_factory)
    service = ExchangeService(session_factory=session_factory)
    xlsx_path = tmp_path / "import.xlsx"

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "patients"
    ws.append(["id", "full_name", "dob", "sex", "category", "military_unit", "military_district"])
    ws.append([1, "Сидоров Сидор", "2024-01-01", "M", "cat", "", ""])
    ws.append([2, "Смирнов Семен", "bad-date", "M", "cat", "", ""])
    wb.save(xlsx_path)

    result = service.import_excel(xlsx_path, actor_id=actor_id, mode="merge")

    assert result["counts"]["patients"] == 2
    assert result["details"]["patients"]["added"] == 1
    assert result["details"]["patients"]["errors"] == 1
    assert result["summary"]["errors"] == 1

    error_log_path = Path(str(result["error_log_path"]))
    assert error_log_path.exists()
    payload = json.loads(error_log_path.read_text(encoding="utf-8"))
    assert payload["errors_count"] == 1
    assert payload["errors"][0]["scope"] == "patients"

    with session_factory() as session:
        patients = session.query(models.Patient).all()
    assert len(patients) == 1
    assert str(patients[0].full_name) == "Сидоров Сидор"


def test_import_zip_returns_nested_import_error_report(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "exchange_zip_report.db")
    actor_id = seed_actor(session_factory)
    service = ExchangeService(session_factory=session_factory)
    xlsx_path = tmp_path / "export.xlsx"
    zip_path = tmp_path / "import.zip"

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "patients"
    ws.append(["id", "full_name", "dob", "sex", "category", "military_unit", "military_district"])
    ws.append([1, "Кузнецов Кирилл", "2024-01-01", "M", "cat", "", ""])
    ws.append([2, "Федоров Федор", "bad-date", "M", "cat", "", ""])
    wb.save(xlsx_path)

    manifest = {
        "schema_version": "1.0",
        "files": [{"name": "export.xlsx", "sha256": sha256_file(xlsx_path)}],
    }
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(xlsx_path, arcname="export.xlsx")
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))

    result = service.import_zip(zip_path, actor_id=actor_id, mode="merge")

    assert result["error_count"] == 1
    assert result["summary"]["errors"] == 1
    error_log_path = Path(str(result["error_log_path"]))
    assert error_log_path.exists()
    payload = json.loads(error_log_path.read_text(encoding="utf-8"))
    assert payload["source_file"] == str(zip_path)
    assert payload["errors_count"] == 1


def test_export_excel_uses_russian_titles_and_logs_package(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "exchange_excel_export.db")
    actor_id = seed_actor(session_factory)
    service = ExchangeService(session_factory=session_factory)
    xlsx_path = tmp_path / "export.xlsx"

    with session_factory() as session:
        session.add(
            models.Patient(
                full_name="Иванов Иван Иванович",
                dob=date(1997, 1, 24),
                sex="M",
                category="офицер",
                military_unit="1 рота",
                military_district="ЦВО",
            )
        )

    result = service.export_excel(xlsx_path, exported_by="exchange_admin", actor_id=actor_id)

    assert Path(result["path"]).exists()
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    assert "Пациенты" in workbook.sheetnames
    assert "patients" not in workbook.sheetnames
    patient_sheet = workbook["Пациенты"]
    headers = next(patient_sheet.iter_rows(values_only=True))
    assert headers is not None
    assert list(headers[:4]) == ["ID пациента", "ФИО", "Дата рождения", "Пол"]

    with session_factory() as session:
        packages = session.query(models.DataExchangePackage).all()
    assert len(packages) == 1
    assert packages[0].direction == "export"
    assert packages[0].package_format == "excel"


def test_import_excel_accepts_russian_headers_and_logs_package(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "exchange_excel_russian_import.db")
    actor_id = seed_actor(session_factory)
    service = ExchangeService(session_factory=session_factory)
    xlsx_path = tmp_path / "import_russian.xlsx"

    wb = Workbook()
    meta = wb.active
    assert meta is not None
    meta.title = "meta"
    ws = wb.create_sheet("Пациенты")
    ws.append(
        [
            "ID пациента",
            "ФИО",
            "Дата рождения",
            "Пол",
            "Категория",
            "Воинская часть",
            "Военный округ",
            "Создано",
        ]
    )
    ws.append([1, "Сидоров Сидор", "24.01.1997", "M", "офицер", "2 рота", "ЮВО", "02.03.2026 13:12"])
    wb.save(xlsx_path)

    result = service.import_excel(xlsx_path, actor_id=actor_id, mode="merge")

    assert result["summary"]["added"] == 1
    assert result["summary"]["errors"] == 0
    with session_factory() as session:
        patients = session.query(models.Patient).all()
        packages = session.query(models.DataExchangePackage).all()
    assert len(patients) == 1
    assert str(patients[0].full_name) == "Сидоров Сидор"
    assert len(packages) == 1
    assert packages[0].direction == "import"
    assert packages[0].package_format == "excel"


def test_export_csv_and_pdf_log_packages(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "exchange_flat_exports.db")
    actor_id = seed_actor(session_factory)
    service = ExchangeService(session_factory=session_factory)
    csv_path = tmp_path / "patients.csv"
    pdf_path = tmp_path / "patients.pdf"

    with session_factory() as session:
        session.add(
            models.Patient(
                full_name="Петров Пётр Петрович",
                dob=date(1990, 5, 12),
                sex="M",
                category="сержант",
                military_unit="3 батальон",
                military_district="ЗВО",
            )
        )

    service.export_csv(csv_path, "patients", actor_id=actor_id)
    service.export_pdf(pdf_path, "patients", actor_id=actor_id)

    with session_factory() as session:
        packages = (
            session.query(models.DataExchangePackage)
            .order_by(models.DataExchangePackage.created_at.asc())
            .all()
        )
    assert [pkg.package_format for pkg in packages] == ["csv", "pdf"]


def test_export_excel_to_same_path_creates_complete_history_rows(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "exchange_excel_same_path.db")
    actor_id = seed_actor(session_factory)
    service = ExchangeService(session_factory=session_factory)
    xlsx_path = tmp_path / "export.xlsx"

    with session_factory() as session:
        session.add(
            models.Patient(
                full_name="Иванов Иван Иванович",
                dob=date(1997, 1, 24),
                sex="M",
                category="офицер",
                military_unit="1 рота",
                military_district="ЦВО",
            )
        )

    service.export_excel(xlsx_path, exported_by="exchange_admin", actor_id=actor_id)
    service.export_excel(xlsx_path, exported_by="exchange_admin", actor_id=actor_id)

    with session_factory() as session:
        packages = (
            session.query(models.DataExchangePackage)
            .order_by(models.DataExchangePackage.id.asc())
            .all()
        )

    assert len(packages) == 2
    for package in packages:
        assert package.direction == "export"
        assert package.package_format == "excel"
        assert package.file_path == str(xlsx_path)
        assert package.sha256
        assert package.created_by == actor_id
        assert package.created_at is not None


def test_export_excel_formats_columns_and_bool_values_for_medical_reading(tmp_path: Path) -> None:
    session_factory = make_session_factory(tmp_path / "exchange_excel_formatting.db")
    actor_id = seed_actor(session_factory)
    service = ExchangeService(session_factory=session_factory)
    xlsx_path = tmp_path / "formatted_export.xlsx"

    with session_factory() as session:
        patient = models.Patient(
            full_name="Очень длинное имя пациента для проверки автоподбора ширины столбца",
            dob=date(1997, 1, 24),
            sex="M",
            category="офицер",
            military_unit="Очень длинное подразделение для проверки переноса текста",
            military_district="ЦВО",
        )
        session.add(patient)
        session.flush()

        emr_case = models.EmrCase(
            patient_id=int(patient.id),
            hospital_case_no="EMR-001",
            created_by=actor_id,
        )
        session.add(emr_case)
        session.flush()

        session.add(
            models.EmrCaseVersion(
                emr_case_id=int(emr_case.id),
                version_no=1,
                valid_from=datetime(2026, 4, 16, 10, 0, tzinfo=UTC),
                is_current=True,
                entered_by=actor_id,
            )
        )

    service.export_excel(xlsx_path, exported_by="exchange_admin", actor_id=actor_id)

    workbook = load_workbook(xlsx_path)
    patient_sheet = workbook["Пациенты"]
    version_sheet = workbook["Версии ЭМЗ"]

    assert patient_sheet.column_dimensions["B"].width is not None
    assert patient_sheet.column_dimensions["B"].width > 13
    assert patient_sheet["B2"].alignment.wrap_text is True
    assert patient_sheet["B2"].alignment.vertical == "top"

    headers = next(version_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert headers is not None
    current_column_index = list(headers).index("Текущая версия") + 1
    assert version_sheet.cell(row=2, column=current_column_index).value == "Да"
