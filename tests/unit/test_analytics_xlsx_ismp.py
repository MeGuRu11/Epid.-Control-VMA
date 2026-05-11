from __future__ import annotations

from collections.abc import Callable, Iterator
from contextlib import AbstractContextManager, contextmanager
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.application.dto.analytics_dto import AnalyticsSearchRequest
from app.application.services import reporting_service as reporting_service_module
from app.application.services.reporting_service import ReportingService
from app.infrastructure.db.models_sqlalchemy import Base


class _AnalyticsServiceStub:
    def __init__(self, *, ismp: dict[str, Any]) -> None:
        self.ismp = ismp

    def search_samples(self, _request: AnalyticsSearchRequest) -> list[Any]:
        return []

    def get_aggregates(self, _request: AnalyticsSearchRequest) -> dict[str, Any]:
        return {"total": 0, "positives": 0, "positive_share": 0.0}

    def get_ismp_metrics(
        self,
        *,
        date_from: object,
        date_to: object,
        department_id: object,
    ) -> dict[str, Any]:
        return self.ismp


def _make_session_factory(db_path: Path) -> Callable[[], AbstractContextManager[Session]]:
    engine = create_engine(f"sqlite:///{db_path.as_posix()}", future=True)
    Base.metadata.create_all(engine)
    session_local = sessionmaker(
        bind=engine,
        autoflush=False,
        autocommit=False,
        expire_on_commit=False,
        future=True,
    )

    @contextmanager
    def _session_scope() -> Iterator[Session]:
        session: Session = session_local()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    return _session_scope


def _export_xlsx(tmp_path: Path, ismp: dict[str, Any]) -> Path:
    session_factory = _make_session_factory(tmp_path / "analytics_xlsx_ismp.db")
    reporting_service_module.REPORT_ARTIFACT_DIR = tmp_path / "artifacts"
    service = ReportingService(
        analytics_service=cast(Any, _AnalyticsServiceStub(ismp=ismp)),
        session_factory=session_factory,
    )
    export_path = tmp_path / "analytics.xlsx"
    service.export_analytics_xlsx(AnalyticsSearchRequest(), export_path, actor_id=None)
    return export_path


def _ismp_payload(*, ismp_cases: int = 2) -> dict[str, Any]:
    return {
        "total_cases": 4,
        "total_patient_days": 20,
        "ismp_total": 3 if ismp_cases else 0,
        "ismp_cases": ismp_cases,
        "incidence": 500.0 if ismp_cases else 0.0,
        "incidence_density": 150.0 if ismp_cases else 0.0,
        "prevalence": 50.0 if ismp_cases else 0.0,
        "by_type": [
            {"type": "ВАП", "count": 2},
            {"type": "КА-ИК", "count": 1},
        ]
        if ismp_cases
        else [],
    }


def test_analytics_xlsx_has_ismp_sheet(tmp_path: Path) -> None:
    export_path = _export_xlsx(tmp_path, _ismp_payload())

    workbook = load_workbook(export_path, data_only=True)

    assert "ИСМП" in workbook.sheetnames


def test_analytics_xlsx_ismp_prevalence_cell_is_percent_format(tmp_path: Path) -> None:
    export_path = _export_xlsx(tmp_path, _ismp_payload())

    workbook = load_workbook(export_path, data_only=True)
    sheet = workbook["ИСМП"]

    assert sheet["B6"].value == 0.5
    assert sheet["B6"].number_format == "0.0%"


def test_analytics_xlsx_ismp_metrics_are_numeric_not_string(tmp_path: Path) -> None:
    export_path = _export_xlsx(tmp_path, _ismp_payload())

    workbook = load_workbook(export_path, data_only=True)
    sheet = workbook["ИСМП"]

    assert isinstance(sheet["B2"].value, int)
    assert isinstance(sheet["B4"].value, int | float)
    assert not isinstance(sheet["B4"].value, str)
    assert isinstance(sheet["B5"].value, int | float)
    assert not isinstance(sheet["B5"].value, str)


def test_analytics_xlsx_ismp_zero_cases_shows_placeholder_text(tmp_path: Path) -> None:
    export_path = _export_xlsx(tmp_path, _ismp_payload(ismp_cases=0))

    workbook = load_workbook(export_path, data_only=True)
    sheet = workbook["ИСМП"]
    values = [cell for row in sheet.iter_rows(values_only=True) for cell in row if cell]

    assert "Случаев ИСМП в выбранном периоде не зарегистрировано" in values
