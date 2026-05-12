from __future__ import annotations

import csv
import json
import os
import shutil
import tempfile
import zipfile
from collections.abc import Callable, Iterator
from contextlib import contextmanager, suppress
from datetime import UTC, date, datetime
from pathlib import Path, PurePosixPath
from typing import Literal, cast
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
from sqlalchemy import Boolean, Date, DateTime
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.application.dto.exchange_dto import (
    CsvExportResult,
    CsvImportResult,
    ExcelExportResult,
    ExcelImportResult,
    ExchangeImportErrorEntry,
    ExchangeImportSummary,
    ExchangeManifest,
    ExchangeManifestFileEntry,
    ExchangeTableStats,
    Form100ExchangeService,
    LegacyJsonExportResult,
    LegacyJsonImportResult,
    ZipExportResult,
    ZipImportResult,
)
from app.application.reporting.formatters import to_iso_utc
from app.application.reporting.id_resolver import IdResolver
from app.application.security.role_matrix import Role, has_permission
from app.config import DATA_DIR
from app.domain.types import JSONDict, JSONValue
from app.infrastructure.db import models_sqlalchemy as models
from app.infrastructure.db.repositories.audit_repo import AuditLogRepository
from app.infrastructure.db.repositories.user_repo import UserRepository
from app.infrastructure.db.session import session_scope
from app.infrastructure.reporting.pdf_determinism import build_invariant_pdf
from app.infrastructure.reporting.pdf_fonts import get_pdf_unicode_font_name
from app.infrastructure.security.sha256 import sha256_file

TABLE_MODELS: dict[str, type[models.Base]] = {
    "departments": models.Department,
    "ref_icd10": models.RefICD10,
    "ref_microorganisms": models.RefMicroorganism,
    "ref_antibiotic_groups": models.RefAntibioticGroup,
    "ref_antibiotics": models.RefAntibiotic,
    "ref_phages": models.RefPhage,
    "ref_material_types": models.RefMaterialType,
    "ref_ismp_abbreviations": models.RefIsmpAbbreviation,
    "patients": models.Patient,
    "emr_case": models.EmrCase,
    "ismp_case": models.IsmpCase,
    "emr_case_version": models.EmrCaseVersion,
    "emr_diagnosis": models.EmrDiagnosis,
    "emr_intervention": models.EmrIntervention,
    "emr_antibiotic_course": models.EmrAntibioticCourse,
    "lab_sample": models.LabSample,
    "lab_microbe_isolation": models.LabMicrobeIsolation,
    "lab_abx_susceptibility": models.LabAbxSusceptibility,
    "lab_phage_panel_result": models.LabPhagePanelResult,
    "sanitary_sample": models.SanitarySample,
    "san_microbe_isolation": models.SanMicrobeIsolation,
    "san_abx_susceptibility": models.SanAbxSusceptibility,
    "san_phage_panel_result": models.SanPhagePanelResult,
    "form100": models.Form100V2,
    "form100_data": models.Form100DataV2,
}

CSV_TABLES: dict[str, type[models.Base]] = {
    "lab_sample": models.LabSample,
    "sanitary_sample": models.SanitarySample,
    "patients": models.Patient,
    "emr_case": models.EmrCase,
}

CSV_HEADERS: dict[str, dict[str, str]] = {
    "patients": {
        "id": "ID пациента",
        "full_name": "ФИО",
        "dob": "Дата рождения",
        "sex": "Пол",
        "category": "Категория",
        "military_unit": "Воинская часть",
        "military_district": "Военный округ",
        "created_at": "Создано",
    },
    "emr_case": {
        "id": "ID госпитализации",
        "patient_id": "ID пациента",
        "hospital_case_no": "№ госпитализации",
        "department_id": "ID отделения",
        "created_at": "Создано",
        "created_by": "Создал (ID)",
        "created_by_name": "Создал",
    },
    "lab_sample": {
        "id": "ID пробы",
        "patient_id": "ID пациента",
        "emr_case_id": "ID госпитализации",
        "lab_no": "Лаб. номер",
        "barcode": "Штрихкод",
        "material_type_id": "ID типа материала",
        "material_type_name": "Тип материала",
        "material_location": "Локализация материала",
        "medium": "Среда",
        "study_kind": "Тип исследования",
        "ordered_at": "Назначено",
        "taken_at": "Взято",
        "delivered_at": "Доставлено",
        "growth_result_at": "Результат роста",
        "growth_flag": "Рост",
        "colony_desc": "Колонии/морфология",
        "microscopy": "Микроскопия",
        "cfu": "КОЕ",
        "qc_due_at": "Срок QC",
        "qc_status": "Статус QC",
        "created_at": "Создано",
        "created_by": "Создал (ID)",
        "created_by_name": "Создал",
    },
    "sanitary_sample": {
        "id": "ID пробы",
        "department_id": "ID отделения",
        "department_name": "Отделение",
        "room": "Помещение",
        "sampling_point": "Точка отбора",
        "lab_no": "Лаб. номер",
        "barcode": "Штрихкод",
        "medium": "Среда",
        "ordered_at": "Назначено",
        "taken_at": "Взято",
        "delivered_at": "Доставлено",
        "growth_result_at": "Результат роста",
        "growth_flag": "Рост",
        "colony_desc": "Колонии/морфология",
        "microscopy": "Микроскопия",
        "cfu": "КОЕ",
        "created_at": "Создано",
        "created_by": "Создал (ID)",
        "created_by_name": "Создал",
    },
}

EXCEL_SHEET_TITLES: dict[str, str] = {
    "departments": "Отделения",
    "ref_icd10": "МКБ-10",
    "ref_microorganisms": "Микроорганизмы",
    "ref_antibiotic_groups": "Группы антибиотиков",
    "ref_antibiotics": "Антибиотики",
    "ref_phages": "Бактериофаги",
    "ref_material_types": "Типы материала",
    "ref_ismp_abbreviations": "ИСМП сокращения",
    "patients": "Пациенты",
    "emr_case": "Госпитализации",
    "ismp_case": "ИСМП случаи",
    "emr_case_version": "Версии ЭМЗ",
    "emr_diagnosis": "Диагнозы ЭМЗ",
    "emr_intervention": "Вмешательства ЭМЗ",
    "emr_antibiotic_course": "Курсы антибиотиков",
    "lab_sample": "Лабораторные пробы",
    "lab_microbe_isolation": "Лаб. выделения",
    "lab_abx_susceptibility": "Лаб. чувствительность",
    "lab_phage_panel_result": "Лаб. фаги",
    "sanitary_sample": "Санитарные пробы",
    "san_microbe_isolation": "Сан. выделения",
    "san_abx_susceptibility": "Сан. чувствительность",
    "san_phage_panel_result": "Сан. фаги",
    "form100": "Форма 100",
    "form100_data": "Форма 100 данные",
}

EXCEL_COLUMN_HEADERS: dict[str, dict[str, str]] = {
    "departments": {
        "id": "ID отделения",
        "name": "Наименование отделения",
    },
    "ref_icd10": {
        "code": "Код МКБ-10",
        "title": "Наименование диагноза",
        "is_active": "Активна",
    },
    "ref_microorganisms": {
        "id": "ID микроорганизма",
        "code": "Код",
        "name": "Микроорганизм",
        "taxon_group": "Таксономическая группа",
        "is_active": "Активен",
    },
    "ref_antibiotic_groups": {
        "id": "ID группы",
        "code": "Код группы",
        "name": "Группа антибиотиков",
    },
    "ref_antibiotics": {
        "id": "ID антибиотика",
        "code": "Код антибиотика",
        "name": "Антибиотик",
        "group_id": "ID группы антибиотиков",
    },
    "ref_phages": {
        "id": "ID фага",
        "code": "Код фага",
        "name": "Бактериофаг",
        "is_active": "Активен",
    },
    "ref_material_types": {
        "id": "ID типа материала",
        "code": "Код типа материала",
        "name": "Тип материала",
    },
    "patients": CSV_HEADERS["patients"],
    "emr_case": CSV_HEADERS["emr_case"],
    "emr_case_version": {
        "id": "ID версии ЭМЗ",
        "emr_case_id": "ID госпитализации",
        "version_no": "Версия",
        "valid_from": "Действует с",
        "valid_to": "Действует по",
        "is_current": "Текущая версия",
        "entered_by": "Кто внёс",
        "admission_date": "Дата поступления",
        "injury_date": "Дата травмы",
        "outcome_date": "Дата исхода",
        "outcome_type": "Тип исхода",
        "severity": "Тяжесть",
        "vph_sp_score": "ВПХ-СП",
        "vph_p_or_score": "ВПХ-П/ОР",
        "sofa_score": "SOFA",
        "days_to_admission": "Дней до поступления",
        "length_of_stay_days": "Длительность пребывания, дней",
    },
    "emr_diagnosis": {
        "id": "ID диагноза",
        "emr_case_version_id": "ID версии ЭМЗ",
        "kind": "Тип диагноза",
        "icd10_code": "Код МКБ-10",
        "free_text": "Свободное описание",
    },
    "emr_intervention": {
        "id": "ID вмешательства",
        "emr_case_version_id": "ID версии ЭМЗ",
        "type": "Тип вмешательства",
        "start_dt": "Начало",
        "end_dt": "Окончание",
        "duration_minutes": "Длительность, мин",
        "performed_by": "Исполнитель",
        "notes": "Примечания",
    },
    "emr_antibiotic_course": {
        "id": "ID курса",
        "emr_case_version_id": "ID версии ЭМЗ",
        "start_dt": "Начало курса",
        "end_dt": "Окончание курса",
        "antibiotic_id": "ID антибиотика",
        "drug_name_free": "Антибиотик (свободно)",
        "route": "Путь введения",
        "dose": "Доза",
    },
    "lab_sample": CSV_HEADERS["lab_sample"],
    "lab_microbe_isolation": {
        "id": "ID выделения",
        "lab_sample_id": "ID лабораторной пробы",
        "microorganism_id": "ID микроорганизма",
        "microorganism_free": "Микроорганизм (свободно)",
        "notes": "Примечания",
    },
    "lab_abx_susceptibility": {
        "id": "ID результата чувствительности",
        "lab_sample_id": "ID лабораторной пробы",
        "antibiotic_id": "ID антибиотика",
        "group_id": "ID группы антибиотиков",
        "ris": "RIS",
        "mic_mg_l": "MIC, мг/л",
        "method": "Метод",
    },
    "lab_phage_panel_result": {
        "id": "ID результата фагов",
        "lab_sample_id": "ID лабораторной пробы",
        "phage_id": "ID фага",
        "phage_free": "Фаг (свободно)",
        "lysis_diameter_mm": "Диаметр лизиса, мм",
    },
    "sanitary_sample": CSV_HEADERS["sanitary_sample"],
    "san_microbe_isolation": {
        "id": "ID выделения",
        "sanitary_sample_id": "ID санитарной пробы",
        "microorganism_id": "ID микроорганизма",
        "microorganism_free": "Микроорганизм (свободно)",
        "notes": "Примечания",
    },
    "san_abx_susceptibility": {
        "id": "ID результата чувствительности",
        "sanitary_sample_id": "ID санитарной пробы",
        "antibiotic_id": "ID антибиотика",
        "group_id": "ID группы антибиотиков",
        "ris": "RIS",
        "mic_mg_l": "MIC, мг/л",
        "method": "Метод",
    },
    "san_phage_panel_result": {
        "id": "ID результата фагов",
        "sanitary_sample_id": "ID санитарной пробы",
        "phage_id": "ID фага",
        "phage_free": "Фаг (свободно)",
        "lysis_diameter_mm": "Диаметр лизиса, мм",
    },
}

_EXCEL_TITLE_TO_TABLE = {title: table_name for table_name, title in EXCEL_SHEET_TITLES.items()}

_HANDLED_IMPORT_ERRORS = (
    ValueError,
    TypeError,
    KeyError,
    AttributeError,
    IndexError,
    SQLAlchemyError,
)

_EXPORT_BATCH_SIZE = 500
_FORM100_PDF_EXPORT_NOTE = "PDF-артефакты карточек Формы 100 экспортируются отдельно через Form100 ZIP"
_FULL_EXPORT_NOTES: dict[str, str] = {"form100_pdf": _FORM100_PDF_EXPORT_NOTE}
_JSON_LIST_COLUMNS = frozenset(
    {
        "bodymap_annotations_json",
        "bodymap_tissue_types_json",
        "features_json",
        "trauma_types_json",
        "wound_types_json",
    }
)
_EXPORT_PATH_COLUMNS = frozenset({"artifact_path", "file_path", "path"})
_CSV_EXTRA_COLUMNS_KEY = "__extra_columns__"
_CSV_ENUM_VALUES: dict[str, dict[str, set[str]]] = {
    "patients": {"sex": {"M", "F", "U"}},
}
_DATE_FORMAT_HINT = "Ожидается формат YYYY-MM-DD или ДД.ММ.ГГГГ."


def _format_date_value(value: object) -> JSONValue | object:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return value


def _format_bool_value(value: bool) -> str:
    return "Да" if value else "Нет"


def _serialize_value(value: object) -> JSONValue | object:
    if isinstance(value, bool):
        return _format_bool_value(value)
    return _format_date_value(value)


def _serialize_json_value(value: object) -> JSONValue | object:
    if isinstance(value, datetime):
        return to_iso_utc(value)
    if isinstance(value, date):
        return value.isoformat()
    return value


def _is_json_column(column_name: str) -> bool:
    return column_name.endswith("_json")


def _json_column_default(column_name: str) -> JSONValue:
    return [] if column_name in _JSON_LIST_COLUMNS else {}


def _parse_json_payload(value: object, *, column_name: str) -> JSONValue:
    if value is None or value == "":
        return _json_column_default(column_name)
    if isinstance(value, dict | list):
        return cast(JSONValue, value)
    try:
        return cast(JSONValue, json.loads(str(value)))
    except (TypeError, ValueError, json.JSONDecodeError):
        return _json_column_default(column_name)


def _portable_export_path(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return text
    normalized = text.replace("\\", "/")
    posix_path = PurePosixPath(normalized)
    path = Path(text)
    if path.is_absolute() or posix_path.is_absolute() or (posix_path.parts and ":" in posix_path.parts[0]):
        for base_dir in (DATA_DIR, Path.cwd()):
            try:
                return path.resolve(strict=False).relative_to(base_dir.resolve(strict=False)).as_posix()
            except (OSError, RuntimeError, ValueError):
                continue
        return posix_path.name
    if ".." in posix_path.parts:
        return posix_path.name
    return posix_path.as_posix()


def _model_to_dict(
    obj: models.Base,
    *,
    nested_json: bool = False,
    portable_paths: bool = False,
    machine_json: bool = False,
) -> JSONDict:
    data: JSONDict = {}
    for column in obj.__table__.columns:
        column_name = str(column.name)
        value = getattr(obj, column_name)
        if nested_json and _is_json_column(column_name):
            data[column_name] = _parse_json_payload(value, column_name=column_name)
        elif portable_paths and column_name in _EXPORT_PATH_COLUMNS:
            data[column_name] = cast(JSONValue, _portable_export_path(value))
        elif machine_json:
            data[column_name] = cast(JSONValue, _serialize_json_value(value))
        else:
            data[column_name] = cast(JSONValue, _serialize_value(value))
    return data


def _get_csv_headers(table_name: str, columns: list[str]) -> list[str]:
    header_map = CSV_HEADERS.get(table_name, {})
    return [header_map.get(col, col) for col in columns]


def _build_extended_columns(table_name: str, columns: list[str]) -> list[str]:
    resolved_after: dict[str, dict[str, str]] = {
        "lab_sample": {"material_type_id": "material_type_name"},
        "sanitary_sample": {"department_id": "department_name"},
    }
    table_resolved_after = resolved_after.get(table_name, {})
    extended_columns: list[str] = []
    for column in columns:
        extended_columns.append(column)
        resolved_column = table_resolved_after.get(column)
        if resolved_column and resolved_column not in extended_columns:
            extended_columns.append(resolved_column)
        if column == "created_by" and "created_by_name" not in extended_columns:
            extended_columns.append("created_by_name")
    return extended_columns


def _coerce_int_id(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    try:
        return int(str(value))
    except ValueError:
        return None


def _fill_resolved_fields(record: JSONDict, table_name: str, resolver: IdResolver) -> None:
    """Добавить в record резолвленные имена для FK-колонок."""
    if table_name == "lab_sample":
        record["material_type_name"] = resolver.resolve_material_type(
            _coerce_int_id(record.get("material_type_id"))
        )
    if table_name == "sanitary_sample":
        record["department_name"] = resolver.resolve_department(
            _coerce_int_id(record.get("department_id"))
        )
    if "created_by" in record:
        record["created_by_name"] = resolver.resolve_user(_coerce_int_id(record.get("created_by")))


def _map_csv_header_name(table_name: str, header_name: object) -> str:
    header_map = CSV_HEADERS.get(table_name, {})
    reverse_map = {label: key for key, label in header_map.items()}
    normalized = str(header_name).lstrip("\ufeff").strip()
    return reverse_map.get(normalized, normalized)


def _map_csv_row(table_name: str, row: dict[str, object]) -> dict[str, object]:
    mapped: dict[str, object] = {}
    for key, value in row.items():
        if key == _CSV_EXTRA_COLUMNS_KEY:
            continue
        mapped[_map_csv_header_name(table_name, key)] = value
    return mapped


def _string_value(value: object) -> str | None:
    if value is None:
        return None
    return str(value)


def _make_import_error(
    *,
    scope: str,
    row: int,
    field: str | None,
    value: object,
    error_code: str,
    message: str,
    hint: str | None = None,
) -> ExchangeImportErrorEntry:
    return {
        "scope": scope,
        "row": row,
        "field": field,
        "value": _string_value(value),
        "error_code": error_code,
        "message": message,
        "hint": hint,
    }


def _csv_row_shape_errors(
    *,
    table_name: str,
    row_idx: int,
    row: dict[str, object],
) -> list[ExchangeImportErrorEntry]:
    extra_values = row.get(_CSV_EXTRA_COLUMNS_KEY)
    if extra_values:
        return [
            _make_import_error(
                scope=table_name,
                row=row_idx,
                field=None,
                value=extra_values,
                error_code="extra_columns",
                message=f"Строка {row_idx}: больше значений, чем заголовков.",
                hint="Проверьте разделители и количество колонок в CSV.",
            )
        ]
    missing_headers = [str(key) for key, value in row.items() if key != _CSV_EXTRA_COLUMNS_KEY and value is None]
    if missing_headers:
        return [
            _make_import_error(
                scope=table_name,
                row=row_idx,
                field=", ".join(missing_headers),
                value=None,
                error_code="row_length_mismatch",
                message=f"Строка {row_idx}: меньше значений, чем заголовков.",
                hint="Проверьте разделители и количество колонок в CSV.",
            )
        ]
    return []


def _validate_csv_row(
    *,
    table_name: str,
    model_cls: type[models.Base],
    row_idx: int,
    row: dict[str, object],
    seen_identities: set[object],
) -> tuple[dict[str, object] | None, list[ExchangeImportErrorEntry]]:
    shape_errors = _csv_row_shape_errors(table_name=table_name, row_idx=row_idx, row=row)
    if shape_errors:
        return None, shape_errors

    mapped_row = _map_csv_row(table_name, row)
    errors: list[ExchangeImportErrorEntry] = []
    pk_cols = list(model_cls.__mapper__.primary_key)
    if len(pk_cols) == 1:
        pk_col = pk_cols[0]
        pk_name = str(pk_col.name)
        pk_value = mapped_row.get(pk_name)
        if pk_value is None or str(pk_value).strip() == "":
            errors.append(
                _make_import_error(
                    scope=table_name,
                    row=row_idx,
                    field=pk_name,
                    value=pk_value,
                    error_code="missing_required",
                    message=f"Строка {row_idx}, поле «{pk_name}»: обязательное значение отсутствует.",
                    hint="Укажите идентификатор строки для CSV-импорта.",
                )
            )
        else:
            try:
                identity = _parse_value(pk_value, pk_col)
            except (TypeError, ValueError, IndexError) as exc:
                errors.append(
                    _make_import_error(
                        scope=table_name,
                        row=row_idx,
                        field=pk_name,
                        value=pk_value,
                        error_code="invalid_value",
                        message=f"Строка {row_idx}, поле «{pk_name}»: некорректное значение.",
                        hint=str(exc) or None,
                    )
                )
            else:
                if identity in seen_identities:
                    errors.append(
                        _make_import_error(
                            scope=table_name,
                            row=row_idx,
                            field=pk_name,
                            value=pk_value,
                            error_code="duplicate_row",
                            message=f"Строка {row_idx}, поле «{pk_name}»: дубликат идентификатора в файле.",
                            hint="Оставьте в файле только одну строку с этим идентификатором.",
                        )
                    )
                else:
                    seen_identities.add(identity)

    enum_fields = _CSV_ENUM_VALUES.get(table_name, {})
    for field_name, allowed_values in enum_fields.items():
        value = mapped_row.get(field_name)
        if value is not None and str(value).strip() and str(value).strip() not in allowed_values:
            errors.append(
                _make_import_error(
                    scope=table_name,
                    row=row_idx,
                    field=field_name,
                    value=value,
                    error_code="invalid_enum_value",
                    message=f"Строка {row_idx}, поле «{field_name}»: недопустимое значение «{value}».",
                    hint=f"Допустимые значения: {', '.join(sorted(allowed_values))}.",
                )
            )

    for column in model_cls.__table__.columns:
        column_name = str(column.name)
        column_type = getattr(column, "type", None)
        value = mapped_row.get(column_name)
        if isinstance(column_type, Date) and value not in (None, ""):
            try:
                _parse_value(value, column)
            except (TypeError, ValueError, IndexError):
                errors.append(
                    _make_import_error(
                        scope=table_name,
                        row=row_idx,
                        field=column_name,
                        value=value,
                        error_code="invalid_date_format",
                        message=f"Строка {row_idx}, поле «{column_name}»: некорректный формат «{value}».",
                        hint=_DATE_FORMAT_HINT,
                    )
                )

    return (None if errors else mapped_row), errors


def _get_excel_sheet_title(table_name: str) -> str:
    return EXCEL_SHEET_TITLES.get(table_name, table_name)


def _resolve_excel_table_name(sheet_name: str) -> str | None:
    if sheet_name == "meta":
        return None
    if sheet_name in TABLE_MODELS:
        return sheet_name
    return _EXCEL_TITLE_TO_TABLE.get(sheet_name)


def _get_excel_headers(table_name: str, columns: list[str]) -> list[str]:
    header_map = EXCEL_COLUMN_HEADERS.get(table_name, {})
    return [header_map.get(column, column) for column in columns]


def _map_excel_row(table_name: str, row: dict[str, object]) -> dict[str, object]:
    return {_map_excel_header_name(table_name, str(key)): value for key, value in row.items()}


def _map_excel_header_name(table_name: str, header_name: str) -> str:
    header_map = EXCEL_COLUMN_HEADERS.get(table_name, {})
    reverse_map = {label: key for key, label in header_map.items()}
    return reverse_map.get(header_name, header_name)


def _parse_value(value: object, column: object) -> object:
    if value is None or value == "":
        return None
    if isinstance(value, date | datetime):
        return value
    value_text = str(value)
    column_type = getattr(column, "type", None)
    if isinstance(column_type, Date):
        try:
            return date.fromisoformat(value_text)
        except (TypeError, ValueError):
            parts = value_text.split(".")
            return date(int(parts[2]), int(parts[1]), int(parts[0]))
    if isinstance(column_type, DateTime):
        try:
            return datetime.fromisoformat(value_text)
        except (TypeError, ValueError):
            try:
                return datetime.strptime(value_text, "%d.%m.%Y %H:%M").replace(tzinfo=UTC)
            except (TypeError, ValueError):
                return datetime.strptime(value_text, "%d.%m.%Y %H:%M:%S").replace(tzinfo=UTC)
    if isinstance(column_type, Boolean):
        normalized = value_text.strip().lower()
        if normalized in {"true", "1", "yes", "да", "y"}:
            return True
        if normalized in {"false", "0", "no", "нет", "n"}:
            return False
    return value


def _format_excel_worksheet(worksheet) -> None:
    min_width = 12
    max_width = 56
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in worksheet.iter_cols():
        if not column_cells:
            continue
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            value_length = len(str(cell.value))
            if value_length > max_length:
                max_length = value_length
        width = max(min_width, min(max_length + 4, max_width))
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _finalize_excel_workbook(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        _format_excel_worksheet(worksheet)


def _prepare_import_value(value: object, column: object) -> object:
    column_name = str(getattr(column, "name", ""))
    if _is_json_column(column_name) and isinstance(value, dict | list):
        return json.dumps(value, ensure_ascii=False, default=str)
    return _parse_value(value, column)


def _dict_to_model(model_cls: type[models.Base], data: dict[str, object]) -> models.Base:
    obj = model_cls()
    for column in model_cls.__table__.columns:
        name = column.name
        if name in data:
            setattr(obj, name, _prepare_import_value(data[name], column))
    return obj


def _iter_model_rows(session: Session, model_cls: type[models.Base]) -> Iterator[models.Base]:
    for row in session.query(model_cls).yield_per(_EXPORT_BATCH_SIZE):
        yield cast(models.Base, row)


def _safe_extract_zip(zip_file: zipfile.ZipFile, destination: Path) -> list[Path]:
    """
    Safely extract ZIP archive contents into destination.

    Protects against path traversal (Zip Slip) by rejecting entries that try to
    escape destination via absolute paths, drive prefixes, or '..' components.
    """
    destination = destination.resolve()
    extracted: list[Path] = []
    for member in zip_file.infolist():
        if member.is_dir():
            continue

        raw_name = member.filename.replace("\\", "/")
        pure_path = PurePosixPath(raw_name)
        if not raw_name:
            raise ValueError(f"Недопустимый путь в архиве (пустое имя): {member.filename}")
        if pure_path.is_absolute():
            raise ValueError(f"Недопустимый путь в архиве (абсолютный путь): {member.filename}")
        if ".." in pure_path.parts:
            raise ValueError(f"Недопустимый путь в архиве (path traversal): {member.filename}")
        if pure_path.parts and ":" in pure_path.parts[0]:
            raise ValueError(f"Недопустимый путь в архиве (префикс диска): {member.filename}")

        target_path = (destination / Path(*pure_path.parts)).resolve()
        try:
            target_path.relative_to(destination)
        except ValueError as exc:
            raise ValueError(f"Недопустимый путь в архиве (выход за каталог): {member.filename}") from exc

        target_path.parent.mkdir(parents=True, exist_ok=True)
        with zip_file.open(member, "r") as src, target_path.open("wb") as dst:
            while True:
                chunk = src.read(1024 * 1024)
                if not chunk:
                    break
                dst.write(chunk)
        extracted.append(target_path)
    return extracted


def _get_pk_identity(model_cls: type[models.Base], data: dict[str, object]) -> object | None:
    pk_cols = list(model_cls.__mapper__.primary_key)
    if len(pk_cols) != 1:
        return None
    pk_name = pk_cols[0].name
    if pk_name not in data or data[pk_name] in (None, ""):
        return None
    return _parse_value(data[pk_name], pk_cols[0])


def _format_import_error(exc: Exception) -> str:
    message = str(exc).strip()
    return message or exc.__class__.__name__


def _write_import_error_log(
    source_file: Path,
    errors: list[ExchangeImportErrorEntry],
    summary: ExchangeImportSummary | None = None,
    *,
    started_at: datetime | None = None,
    finished_at: datetime | None = None,
) -> str | None:
    if not errors:
        return None
    created_at = finished_at or datetime.now(UTC)
    timestamp = created_at.strftime("%Y%m%d_%H%M%S")
    log_path = source_file.with_name(f"{source_file.stem}_import_errors_{timestamp}.json")
    payload_summary: dict[str, object] = {
        "total": int(summary.get("total", 0)) if summary else 0,
        "imported": int(summary.get("imported", 0)) if summary else 0,
        "skipped": int(summary.get("skipped", 0)) if summary else 0,
        "errors": len(errors),
        "started_at": (started_at or created_at).isoformat(),
        "finished_at": created_at.isoformat(),
        "source_file": source_file.name,
        "source_sha256": sha256_file(source_file),
    }
    payload = {
        "created_at": created_at.isoformat(),
        "source_file": str(source_file),
        "source_sha256": payload_summary["source_sha256"],
        "errors_count": len(errors),
        "summary": payload_summary,
        "errors": errors,
    }
    try:
        log_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError:
        return None
    return str(log_path)


def _build_import_summary(
    details: dict[str, ExchangeTableStats],
    *,
    errors_count: int,
) -> ExchangeImportSummary:
    rows_total = 0
    added = 0
    updated = 0
    skipped = 0
    for item in details.values():
        rows_total += int(item.get("rows", 0))
        added += int(item.get("added", 0))
        updated += int(item.get("updated", 0))
        skipped += int(item.get("skipped", 0))
    imported = added + updated
    return {
        "rows_total": rows_total,
        "total": rows_total,
        "imported": imported,
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "errors": errors_count,
    }


def _audit_format(package_format: str) -> str:
    if package_format.startswith("zip"):
        return "zip"
    if package_format.startswith("form100"):
        return "form100_zip"
    return package_format


def _first_error_code(errors: list[ExchangeImportErrorEntry] | None) -> str | None:
    if not errors:
        return None
    first = errors[0]
    code = first.get("error_code")
    return str(code) if code else "import_error"


def _build_exchange_audit_payload(
    *,
    direction: str,
    package_format: str,
    scope_tables: list[str],
    file_sha256: str,
    rows_affected: int,
    errors_count: int,
    first_error_code: str | None,
) -> dict[str, object]:
    if direction == "export":
        action = "data_export"
    else:
        action = "data_import_failed" if errors_count else "data_import"
    error_summary: dict[str, object] | None = None
    if errors_count:
        error_summary = {"errors_count": errors_count, "first_error_code": first_error_code or "import_error"}
    return {
        "schema": "exchange.audit.v1",
        "action": action,
        "direction": direction,
        "format": _audit_format(package_format),
        "package_format": package_format,
        "scope_tables": sorted(dict.fromkeys(scope_tables)),
        "file_sha256": file_sha256,
        "rows_affected": rows_affected,
        "error_summary": error_summary,
    }


@contextmanager
def _working_temp_dir() -> Iterator[Path]:
    """
    Create a writable temporary directory with fallback strategy.

    In some restricted environments, directories created by tempfile may be
    inaccessible for writing. We verify writability and fall back to a
    workspace-local directory when needed.
    """
    roots = [Path(tempfile.gettempdir()), Path.cwd() / "tmp_run"]
    last_error: OSError | None = None

    for root in roots:
        temp_dir = root / f"epid-temp-{uuid4().hex}"
        try:
            root.mkdir(parents=True, exist_ok=True)
            temp_dir.mkdir(parents=True, exist_ok=False)
            probe_file = temp_dir / ".write_probe"
            probe_file.write_text("ok", encoding="utf-8")
            probe_file.unlink(missing_ok=True)
            try:
                yield temp_dir
            finally:
                shutil.rmtree(temp_dir, ignore_errors=True)
            return
        except OSError as exc:
            last_error = exc
            shutil.rmtree(temp_dir, ignore_errors=True)

    raise OSError("Не удалось создать временный каталог для операции импорта/экспорта") from last_error


class ExchangeService:
    def __init__(
        self,
        session_factory: Callable = session_scope,
        form100_v2_service: Form100ExchangeService | None = None,
        user_repo: UserRepository | None = None,
        audit_repo: AuditLogRepository | None = None,
    ) -> None:
        self.session_factory = session_factory
        self.form100_v2_service = form100_v2_service
        self.user_repo = user_repo or UserRepository()
        self.audit_repo = audit_repo or AuditLogRepository()

    def _prepare_output_dir(self, path: Path) -> None:
        path.mkdir(parents=True, exist_ok=True)
        # Best-effort: some filesystems/platforms ignore chmod semantics.
        with suppress(OSError):
            os.chmod(path, 0o700)

    def _require_permission(self, actor_id: int, permission: Literal["manage_exchange"]) -> None:
        with self.session_factory() as session:
            actor = self.user_repo.get_by_id(session, actor_id)
            if actor is None:
                raise ValueError("Пользователь не найден")
            actor_role = cast(Role, str(actor.role))
            if has_permission(actor_role, permission):
                return
        raise ValueError("Недостаточно прав для операций импорта/экспорта")

    def _log_package(
        self,
        direction: str,
        package_format: str,
        file_path: Path,
        sha256: str,
        created_by: int | None,
        *,
        scope_tables: list[str] | None = None,
        rows_affected: int = 0,
        errors_count: int = 0,
        first_error_code: str | None = None,
    ) -> None:
        audit_payload = _build_exchange_audit_payload(
            direction=direction,
            package_format=package_format,
            scope_tables=scope_tables or [],
            file_sha256=sha256,
            rows_affected=rows_affected,
            errors_count=errors_count,
            first_error_code=first_error_code,
        )
        with self.session_factory() as session:
            session.add(
                models.DataExchangePackage(
                    direction=direction,
                    package_format=package_format,
                    file_path=str(file_path),
                    sha256=sha256,
                    created_by=created_by,
                )
            )
            self.audit_repo.add_event(
                session,
                user_id=created_by,
                entity_type="exchange",
                entity_id=package_format,
                action=str(audit_payload["action"]),
                payload_json=json.dumps(audit_payload, ensure_ascii=False),
            )

    def _record_package(
        self,
        direction: str,
        package_format: str,
        file_path: Path,
        created_by: int | None,
        *,
        scope_tables: list[str] | None = None,
        rows_affected: int = 0,
        errors: list[ExchangeImportErrorEntry] | None = None,
    ) -> str:
        package_hash = sha256_file(file_path)
        errors_count = len(errors or [])
        self._log_package(
            direction,
            package_format,
            file_path,
            package_hash,
            created_by,
            scope_tables=scope_tables,
            rows_affected=rows_affected,
            errors_count=errors_count,
            first_error_code=_first_error_code(errors),
        )
        return package_hash

    def get_actor_label(self, actor_id: int | None) -> str:
        if actor_id is None:
            return "—"
        with self.session_factory() as session:
            actor = self.user_repo.get_by_id(session, actor_id)
            if actor is None:
                return str(actor_id)
            login = str(getattr(actor, "login", "") or "").strip()
            return login or str(actor_id)

    def export_excel(
        self,
        file_path: str | Path,
        *,
        exported_by: str | None = None,
        actor_id: int,
        log_package: bool = True,
    ) -> ExcelExportResult:
        self._require_permission(actor_id, "manage_exchange")
        file_path = Path(file_path)
        wb = Workbook()
        meta = wb.active
        if meta is None:
            raise RuntimeError("Не удалось создать лист meta для экспорта")
        meta.title = "meta"
        meta.sheet_state = "hidden"
        meta.append(["schema_version", "1.0"])
        meta.append(["exported_at", datetime.now(UTC).isoformat()])
        meta.append(["exported_by", exported_by or ""])
        meta.append(["note_form100_pdf", _FORM100_PDF_EXPORT_NOTE])

        counts: dict[str, int] = {}
        with self.session_factory() as session:
            for name, model_cls in TABLE_MODELS.items():
                ws = wb.create_sheet(title=_get_excel_sheet_title(name))
                columns = [c.name for c in model_cls.__table__.columns]
                ws.append(_get_excel_headers(name, columns))
                row_count = 0
                for row in _iter_model_rows(session, model_cls):
                    data = _model_to_dict(row)
                    ws.append([data.get(col) for col in columns])
                    row_count += 1
                counts[name] = row_count
        if len(wb.worksheets) > 1:
            wb.active = 1
        _finalize_excel_workbook(wb)

        # TODO SECURITY: добавить шифрование бэкапов/экспортов (AES-GCM)
        self._prepare_output_dir(file_path.parent)
        wb.save(file_path)
        if log_package:
            self._record_package("export", "excel", file_path, actor_id, scope_tables=list(TABLE_MODELS), rows_affected=sum(counts.values()))
        return {"path": str(file_path), "counts": counts}

    def export_zip(self, file_path: str | Path, *, exported_by: str | None = None, actor_id: int) -> ZipExportResult:
        self._require_permission(actor_id, "manage_exchange")
        file_path = Path(file_path)
        with _working_temp_dir() as tmp_dir_path:
            excel_path = tmp_dir_path / "export.xlsx"
            result = self.export_excel(
                excel_path,
                exported_by=exported_by,
                actor_id=actor_id,
                log_package=False,
            )
            files: list[Path] = [excel_path]
            manifest_files: list[ExchangeManifestFileEntry] = []
            manifest: ExchangeManifest = {
                "schema_version": "1.0",
                "exported_at": cast(str, to_iso_utc(datetime.now(UTC))),
                "exported_by": exported_by,
                "files": manifest_files,
                "notes": dict(_FULL_EXPORT_NOTES),
            }
            for f in files:
                manifest_files.append(
                    {
                        "name": f.name,
                        "sha256": sha256_file(f),
                        "size": f.stat().st_size,
                    }
                )
            manifest_path = tmp_dir_path / "manifest.json"
            manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

            self._prepare_output_dir(file_path.parent)
            with zipfile.ZipFile(file_path, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.write(excel_path, arcname=excel_path.name)
                zf.write(manifest_path, arcname=manifest_path.name)

        package_hash = sha256_file(file_path)
        self._log_package("export", "zip+excel", file_path, package_hash, actor_id, scope_tables=list(TABLE_MODELS), rows_affected=sum(result["counts"].values()))
        return {"path": str(file_path), "counts": result["counts"], "sha256": package_hash}

    def import_excel(
        self,
        file_path: str | Path,
        *,
        actor_id: int,
        mode: str = "merge",
        write_error_log: bool = True,
        log_package: bool = True,
    ) -> ExcelImportResult:
        self._require_permission(actor_id, "manage_exchange")
        file_path = Path(file_path)
        wb = load_workbook(file_path, read_only=True, data_only=True)
        counts: dict[str, int] = {}
        details: dict[str, ExchangeTableStats] = {}
        errors: list[ExchangeImportErrorEntry] = []
        with self.session_factory() as session:
            for sheet_name in wb.sheetnames:
                table_name = _resolve_excel_table_name(sheet_name)
                if table_name is None:
                    continue
                model_cls = TABLE_MODELS[table_name]
                ws = wb[sheet_name]
                row_iter = ws.iter_rows(values_only=True)
                header_row = next(row_iter, None)
                if header_row is None:
                    counts[table_name] = 0
                    details[table_name] = {"rows": 0, "added": 0, "updated": 0, "skipped": 0, "errors": 0}
                    continue
                header_positions = [
                    (idx, _map_excel_header_name(table_name, str(val)))
                    for idx, val in enumerate(header_row)
                    if val is not None
                ]
                if not header_positions:
                    counts[table_name] = 0
                    details[table_name] = {"rows": 0, "added": 0, "updated": 0, "skipped": 0, "errors": 0}
                    continue
                rows_total = 0
                added = 0
                updated = 0
                skipped = 0
                sheet_errors = 0
                for row_idx, row in enumerate(row_iter, start=2):
                    rows_total += 1
                    try:
                        data = {
                            header: row[idx] if row is not None and idx < len(row) else None
                            for idx, header in header_positions
                        }
                        identity = _get_pk_identity(model_cls, data)
                        existing = session.get(model_cls, identity) if identity is not None else None
                        if mode == "append" and existing is not None:
                            skipped += 1
                            continue
                        obj = _dict_to_model(model_cls, data)
                        if existing is not None:
                            updated += 1
                        else:
                            added += 1
                        session.merge(obj)
                    except _HANDLED_IMPORT_ERRORS as exc:
                        sheet_errors += 1
                        errors.append(
                            {
                                "scope": table_name,
                                "row": row_idx,
                                "message": _format_import_error(exc),
                            }
                        )
                counts[table_name] = rows_total
                details[table_name] = {
                    "rows": rows_total,
                    "added": added,
                    "updated": updated,
                    "skipped": skipped,
                    "errors": sheet_errors,
                }
        summary = _build_import_summary(details, errors_count=len(errors))
        result: ExcelImportResult = {
            "path": str(file_path),
            "counts": counts,
            "details": details,
            "errors": errors,
            "error_count": len(errors),
            "summary": summary,
        }
        if write_error_log:
            result["error_log_path"] = _write_import_error_log(file_path, errors)
        if log_package:
            self._record_package("import", "excel", file_path, actor_id, scope_tables=list(details), rows_affected=int(summary["imported"]), errors=errors)
        return result

    def import_zip(self, file_path: str | Path, *, actor_id: int, mode: str = "merge") -> ZipImportResult:
        self._require_permission(actor_id, "manage_exchange")
        file_path = Path(file_path)
        with _working_temp_dir() as tmp_dir_path:
            with zipfile.ZipFile(file_path, "r") as zf:
                try:
                    _safe_extract_zip(zf, tmp_dir_path)
                except ValueError as exc:
                    raise ValueError(f"Небезопасный ZIP-архив: {exc}") from exc

            manifest_path = tmp_dir_path / "manifest.json"
            if not manifest_path.exists():
                raise ValueError("В архиве отсутствует manifest.json")
            manifest = cast(ExchangeManifest, json.loads(manifest_path.read_text(encoding="utf-8")))
            manifest_files = manifest.get("files", [])
            for entry in manifest_files:
                f = tmp_dir_path / entry["name"]
                if not f.exists():
                    raise ValueError(f"Файл отсутствует: {entry['name']}")
                if sha256_file(f) != entry.get("sha256"):
                    raise ValueError(f"Хэш не совпадает: {entry['name']}")

            excel_path = tmp_dir_path / "export.xlsx"
            if not excel_path.exists():
                raise ValueError("В архиве отсутствует export.xlsx")
            result = self.import_excel(
                excel_path,
                actor_id=actor_id,
                mode=mode,
                write_error_log=False,
                log_package=False,
            )

        package_hash = sha256_file(file_path)
        errors = result["errors"]
        self._log_package(
            "import",
            "zip+excel",
            file_path,
            package_hash,
            actor_id,
            scope_tables=list(result["details"]),
            rows_affected=int(result["summary"]["imported"]),
            errors_count=len(errors),
            first_error_code=_first_error_code(errors),
        )
        return {
            "path": str(file_path),
            "counts": result["counts"],
            "details": result["details"],
            "errors": errors,
            "error_count": len(errors),
            "error_log_path": _write_import_error_log(file_path, errors),
            "summary": result["summary"],
            "sha256": package_hash,
        }

    def export_form100_package_zip(
        self,
        file_path: str | Path,
        *,
        actor_id: int,
        exported_by: str | None = None,
        card_id: str | None = None,
    ) -> dict[str, object]:
        self._require_permission(actor_id, "manage_exchange")
        if self.form100_v2_service is None:
            raise ValueError("Сервис Form100 не подключён")
        return cast(
            dict[str, object],
            self.form100_v2_service.export_package_zip(
                file_path=file_path,
                actor_id=actor_id,
                card_id=card_id,
                exported_by=exported_by,
            ),
        )

    def import_form100_package_zip(
        self,
        file_path: str | Path,
        *,
        actor_id: int,
        mode: str = "merge",
    ) -> dict[str, object]:
        self._require_permission(actor_id, "manage_exchange")
        if self.form100_v2_service is None:
            raise ValueError("Сервис Form100 не подключён")
        return cast(
            dict[str, object],
            self.form100_v2_service.import_package_zip(
                file_path=file_path,
                actor_id=actor_id,
                mode=mode,
            ),
        )

    def list_packages(
        self, limit: int = 50, direction: str | None = None, query: str | None = None
    ) -> list[models.DataExchangePackage]:
        with self.session_factory() as session:
            q = session.query(models.DataExchangePackage)
            if direction:
                q = q.filter(models.DataExchangePackage.direction == direction)
            if query:
                q = q.filter(
                    models.DataExchangePackage.file_path.ilike(f"%{query}%")
                    | models.DataExchangePackage.sha256.ilike(f"%{query}%")
                )
            rows = q.order_by(models.DataExchangePackage.created_at.desc()).limit(limit).all()
            return cast(list[models.DataExchangePackage], rows)

    def export_csv(self, file_path: str | Path, table_name: str, *, actor_id: int) -> CsvExportResult:
        self._require_permission(actor_id, "manage_exchange")
        file_path = Path(file_path)
        if table_name not in CSV_TABLES:
            raise ValueError("Неизвестная таблица CSV")
        model_cls = CSV_TABLES[table_name]
        with self.session_factory() as session:
            resolver = IdResolver(session)
            columns = [c.name for c in model_cls.__table__.columns]
            extended_columns = _build_extended_columns(table_name, columns)
            self._prepare_output_dir(file_path.parent)
            count = 0
            with file_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(_get_csv_headers(table_name, extended_columns))
                for row in _iter_model_rows(session, model_cls):
                    data = _model_to_dict(row)
                    _fill_resolved_fields(data, table_name, resolver)
                    writer.writerow([data.get(col) for col in extended_columns])
                    count += 1
        self._record_package("export", "csv", file_path, actor_id, scope_tables=[table_name], rows_affected=count)
        return {"path": str(file_path), "count": count}

    def import_csv(
        self, file_path: str | Path, table_name: str, *, actor_id: int, mode: str = "merge"
    ) -> CsvImportResult:
        self._require_permission(actor_id, "manage_exchange")
        file_path = Path(file_path)
        if table_name not in CSV_TABLES:
            raise ValueError("Неизвестная таблица CSV")
        model_cls = CSV_TABLES[table_name]
        started_at = datetime.now(UTC)
        count = 0
        rows_total = 0
        added = updated = skipped = 0
        errors: list[ExchangeImportErrorEntry] = []
        seen_identities: set[object] = set()
        with self.session_factory() as session, file_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, restkey=_CSV_EXTRA_COLUMNS_KEY, restval=None, strict=True)
            try:
                for row_idx, row in enumerate(reader, start=2):
                    rows_total += 1
                    mapped_row, row_errors = _validate_csv_row(
                        table_name=table_name,
                        model_cls=model_cls,
                        row_idx=row_idx,
                        row=cast(dict[str, object], row),
                        seen_identities=seen_identities,
                    )
                    if row_errors:
                        errors.extend(row_errors)
                        skipped += 1
                        continue
                    if mapped_row is None:
                        skipped += 1
                        continue
                    try:
                        identity = _get_pk_identity(model_cls, mapped_row)
                        existing = session.get(model_cls, identity) if identity is not None else None
                        if mode == "append" and existing is not None:
                            skipped += 1
                            continue
                        obj = _dict_to_model(model_cls, mapped_row)
                        with session.begin_nested():
                            session.merge(obj)
                        if existing is not None:
                            updated += 1
                        else:
                            added += 1
                        count += 1
                    except _HANDLED_IMPORT_ERRORS as exc:
                        skipped += 1
                        errors.append(
                            _make_import_error(
                                scope=table_name,
                                row=row_idx,
                                field=None,
                                value=None,
                                error_code="import_row_failed",
                                message=f"Строка {row_idx}: запись не импортирована из-за некорректных данных.",
                                hint=_format_import_error(exc),
                            )
                        )
            except csv.Error as exc:
                errors.append(
                    _make_import_error(
                        scope=table_name,
                        row=rows_total + 2,
                        field=None,
                        value=None,
                        error_code="csv_parse_error",
                        message="CSV-файл повреждён или содержит незакрытые кавычки.",
                        hint=str(exc) or None,
                    )
                )
        details: dict[str, ExchangeTableStats] = {
            table_name: {
                "rows": rows_total,
                "added": added,
                "updated": updated,
                "skipped": skipped,
                "errors": len(errors),
            }
        }
        summary = _build_import_summary(details, errors_count=len(errors))
        finished_at = datetime.now(UTC)
        result: CsvImportResult = {
            "path": str(file_path),
            "count": count,
            "counts": {table_name: rows_total},
            "details": details,
            "errors": errors,
            "error_count": len(errors),
            "error_log_path": _write_import_error_log(
                file_path,
                errors,
                summary,
                started_at=started_at,
                finished_at=finished_at,
            ),
            "summary": summary,
        }
        self._record_package("import", "csv", file_path, actor_id, scope_tables=[table_name], rows_affected=int(summary["imported"]), errors=errors)
        return result

    def export_pdf(self, file_path: str | Path, table_name: str, *, actor_id: int) -> CsvExportResult:
        self._require_permission(actor_id, "manage_exchange")
        file_path = Path(file_path)
        if table_name not in CSV_TABLES:
            raise ValueError("Неизвестная таблица PDF")
        model_cls = CSV_TABLES[table_name]
        unicode_font = get_pdf_unicode_font_name()

        styles = getSampleStyleSheet()
        normal_style = styles["Normal"]
        cell_style = ParagraphStyle(
            "PdfCell",
            parent=normal_style,
            fontName=unicode_font,
            fontSize=6,
            leading=7,
            wordWrap="CJK",
        )

        count = 0
        with self.session_factory() as session:
            resolver = IdResolver(session)
            columns = [c.name for c in model_cls.__table__.columns]
            extended_columns = _build_extended_columns(table_name, columns)
            headers = _get_csv_headers(table_name, extended_columns)
            data = [[Paragraph(h, cell_style) for h in headers]]
            for row in _iter_model_rows(session, model_cls):
                record = _model_to_dict(row)
                _fill_resolved_fields(record, table_name, resolver)
                data.append(
                    [
                        Paragraph("" if record.get(col) is None else str(record.get(col)), cell_style)
                        for col in extended_columns
                    ]
                )
                count += 1

        self._prepare_output_dir(file_path.parent)
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=landscape(A4),
            leftMargin=5 * mm,
            rightMargin=5 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
        )

        # Approximate dynamic widths by equally dividing usable landscape space
        usable_width = landscape(A4)[0] - 10 * mm
        num_cols = len(data[0])
        col_widths = [usable_width / num_cols] * num_cols

        table = Table(data, repeatRows=1, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), unicode_font),
                    ("FONTSIZE", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 1 * mm),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 1 * mm),
                    ("TOPPADDING", (0, 0), (-1, -1), 1 * mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1 * mm),
                ]
            )
        )
        build_invariant_pdf(doc, [table])
        self._record_package("export", "pdf", file_path, actor_id, scope_tables=[table_name], rows_affected=count)
        return {"path": str(file_path), "count": count}

    # Legacy JSON support (not used in UI)
    def export_json(
        self, file_path: str | Path, *, exported_by: str | None = None, actor_id: int
    ) -> LegacyJsonExportResult:
        self._require_permission(actor_id, "manage_exchange")
        file_path = Path(file_path)
        payload: JSONDict = {
            "schema_version": "1.0",
            "exported_at": cast(str, to_iso_utc(datetime.now(UTC))),
            "exported_by": exported_by,
            "data": cast(JSONValue, {}),
            "notes": cast(JSONValue, dict(_FULL_EXPORT_NOTES)),
        }

        with self.session_factory() as session:
            for name, model_cls in TABLE_MODELS.items():
                data_payload = cast(JSONDict, payload["data"])
                data_payload[name] = cast(
                    JSONValue,
                    [
                        _model_to_dict(x, nested_json=True, portable_paths=True, machine_json=True)
                        for x in _iter_model_rows(session, model_cls)
                    ],
                )

        self._prepare_output_dir(file_path.parent)
        file_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        data_payload = cast(JSONDict, payload["data"])
        counts = {k: len(cast(list[object], v)) for k, v in data_payload.items()}
        self._record_package(
            "export",
            "json",
            file_path,
            actor_id,
            scope_tables=list(counts),
            rows_affected=sum(counts.values()),
        )
        return {
            "path": str(file_path),
            "counts": counts,
        }

    def import_json(self, file_path: str | Path, *, actor_id: int, mode: str = "merge") -> LegacyJsonImportResult:
        self._require_permission(actor_id, "manage_exchange")
        file_path = Path(file_path)
        payload = cast(JSONDict, json.loads(file_path.read_text(encoding="utf-8")))
        if payload.get("schema_version") != "1.0":
            raise ValueError("Неподдерживаемая версия схемы")
        data = cast(dict[str, object], payload.get("data") or {})

        counts: dict[str, int] = {}
        details: dict[str, ExchangeTableStats] = {}
        errors: list[ExchangeImportErrorEntry] = []
        with self.session_factory() as session:
            for name, model_cls in TABLE_MODELS.items():
                raw_items = data.get(name)
                items = cast(list[dict[str, object]], raw_items if isinstance(raw_items, list) else [])
                rows_total = 0
                added = updated = skipped = 0
                table_errors = 0
                for idx, item in enumerate(items, start=1):
                    rows_total += 1
                    try:
                        identity = _get_pk_identity(model_cls, item)
                        existing = session.get(model_cls, identity) if identity is not None else None
                        if mode == "append" and existing is not None:
                            skipped += 1
                            continue
                        obj = _dict_to_model(model_cls, item)
                        if existing is not None:
                            updated += 1
                        else:
                            added += 1
                        session.merge(obj)
                    except _HANDLED_IMPORT_ERRORS as exc:
                        table_errors += 1
                        errors.append(
                            {
                                "scope": name,
                                "row": idx,
                                "message": _format_import_error(exc),
                            }
                        )
                counts[name] = rows_total
                details[name] = {
                    "rows": rows_total,
                    "added": added,
                    "updated": updated,
                    "skipped": skipped,
                    "errors": table_errors,
                }
        summary = _build_import_summary(details, errors_count=len(errors))
        self._record_package(
            "import",
            "json",
            file_path,
            actor_id,
            scope_tables=[name for name, stats in details.items() if stats["rows"] or stats["errors"]],
            rows_affected=int(summary["imported"]),
            errors=errors,
        )
        return {
            "path": str(file_path),
            "counts": counts,
            "details": details,
            "errors": errors,
            "error_count": len(errors),
            "error_log_path": _write_import_error_log(file_path, errors),
            "summary": summary,
        }
