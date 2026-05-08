from __future__ import annotations

import json
import logging
import shutil
from collections.abc import Callable
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, cast
from uuid import uuid4

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError

from app.application.dto.analytics_dto import AnalyticsSearchRequest
from app.application.services.analytics_service import AnalyticsService
from app.application.services.form100_service_v2 import Form100ServiceV2
from app.application.services.reference_service import ReferenceService
from app.config import DATA_DIR
from app.domain.constants import MilitaryCategory
from app.infrastructure.db import models_sqlalchemy as models
from app.infrastructure.db.session import session_scope
from app.infrastructure.reporting.pdf_determinism import build_invariant_pdf
from app.infrastructure.reporting.pdf_fonts import get_pdf_unicode_font_name
from app.infrastructure.security.sha256 import sha256_file

REPORT_ARTIFACT_DIR = DATA_DIR / "artifacts" / "reports"

FILTER_LABELS: dict[str, str] = {
    "date_from": "Дата от",
    "date_to": "Дата до",
    "department_id": "Отделение",
    "icd10_code": "МКБ-10",
    "microorganism_id": "Микроорганизм",
    "antibiotic_id": "Антибиотик",
    "material_type_id": "Материал",
    "growth_flag": "Рост",
    "patient_category": "Категория",
    "patient_name": "ФИО пациента",
    "lab_no": "Лаб. номер",
    "search_text": "Поиск",
}

SENSITIVE_FILTER_KEYS = {"patient_name", "fio", "search_text", "lab_no", "passport", "snils"}

logger = logging.getLogger(__name__)

def _format_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, str):
        if "T" in value and len(value) >= 19:
            try:
                dt = datetime.fromisoformat(value)
                return dt.strftime("%d.%m.%Y %H:%M")
            except ValueError:
                pass
        elif len(value) == 10 and "-" in value:
            try:
                d = date.fromisoformat(value)
                return d.strftime("%d.%m.%Y")
            except ValueError:
                pass
    return value

def _format_filter_label(key: str) -> str:
    return FILTER_LABELS.get(key, key)


def _as_int(value: object) -> int:
    return int(cast(Any, value))


class ReportingService:
    def __init__(
        self,
        analytics_service: AnalyticsService,
        form100_v2_service: Form100ServiceV2 | None = None,
        reference_service: ReferenceService | None = None,
        session_factory: Callable = session_scope,
    ) -> None:
        self.analytics_service = analytics_service
        self.form100_v2_service = form100_v2_service
        self.reference_service = reference_service
        self.session_factory = session_factory

    def export_analytics_xlsx(
        self,
        request: AnalyticsSearchRequest,
        file_path: str | Path,
        actor_id: int | None,
    ) -> dict[str, Any]:
        file_path = Path(file_path)
        rows = self.analytics_service.search_samples(request)
        agg = self.analytics_service.get_aggregates(request)

        wb = Workbook()
        summary_ws = wb.active
        if summary_ws is None:
            raise RuntimeError("Не удалось создать лист сводки")
        summary_ws.title = "Сводка"
        summary_ws.append(["Параметр", "Значение"])
        summary_ws.append(["Дата отчета", datetime.now(UTC).strftime("%d.%m.%Y %H:%M")])
        summary_ws.append(["Всего", agg.get("total", 0)])
        summary_ws.append(["Положительные", agg.get("positives", 0)])
        summary_ws.append(["Доля положительных", agg.get("positive_share", 0)])

        filters_ws = wb.create_sheet(title="Фильтры")
        filters_ws.append(["Фильтр", "Значение"])
        filters = request.model_dump(exclude_none=True)
        filter_maps = self._build_filter_maps()
        for key, value in filters.items():
            filters_ws.append(
                [_format_filter_label(key), self._format_filter_value(key, value, filter_maps)]
            )

        data_ws = wb.create_sheet(title="Данные")
        columns = [
            "ID",
            "Лаб. номер",
            "ФИО пациента",
            "Категория",
            "Дата взятия",
            "Отделение",
            "Материал",
            "Микроорганизм",
            "Антибиотик",
        ]
        data_ws.append(columns)
        for row in rows:
            data_ws.append(
                [
                    row.lab_sample_id,
                    row.lab_no,
                    row.patient_name,
                    row.patient_category,
                    _format_value(row.taken_at),
                    row.department_name,
                    row.material_type,
                    row.microorganism,
                    row.antibiotic,
                ]
            )

        file_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)

        artifact_path = self._save_artifact_copy(report_type="analytics", source_path=file_path)
        report_hash = sha256_file(artifact_path)
        report_run_id = self._log_report_run(
            report_type="analytics",
            filters=filters,
            summary=agg,
            file_path=artifact_path,
            sha256=report_hash,
            created_by=actor_id,
        )
        return {
            "path": str(file_path),
            "artifact_path": str(artifact_path),
            "count": len(rows),
            "sha256": report_hash,
            "report_run_id": report_run_id,
        }

    def export_analytics_pdf(
        self,
        request: AnalyticsSearchRequest,
        file_path: str | Path,
        actor_id: int | None,
    ) -> dict[str, Any]:
        file_path = Path(file_path)
        rows = self.analytics_service.search_samples(request)
        agg = self.analytics_service.get_aggregates(request)
        filters = request.model_dump(exclude_none=True)
        filter_maps = self._build_filter_maps()
        unicode_font = get_pdf_unicode_font_name()

        summary_data = [
            ["Параметр", "Значение"],
            ["Дата отчета", datetime.now(UTC).strftime("%d.%m.%Y %H:%M")],
            ["Всего", str(agg.get("total", 0))],
            ["Положительные", str(agg.get("positives", 0))],
            ["Доля положительных", f"{agg.get('positive_share', 0) * 100:.1f}%"],
        ]

        styles = getSampleStyleSheet()
        normal_style = styles["Normal"]
        cell_style = ParagraphStyle(
            "PdfCell",
            parent=normal_style,
            fontName=unicode_font,
            fontSize=7,
            leading=8,
            wordWrap="CJK",
        )

        filter_data: list[list[Paragraph]] = [
            [Paragraph("Фильтр", cell_style), Paragraph("Значение", cell_style)]
        ]
        for key, value in filters.items():
            filter_data.append(
                [
                    Paragraph(_format_filter_label(key), cell_style),
                    Paragraph(self._format_filter_value(key, value, filter_maps), cell_style)
                ]
            )

        headers = [
            "ID",
            "Лаб. номер",
            "ФИО пациента",
            "Категория",
            "Дата взятия",
            "Отделение",
            "Материал",
            "Микроорганизм",
            "Антибиотик",
        ]
        table_data: list[list[Paragraph]] = [[Paragraph(h, cell_style) for h in headers]]
        for row in rows:
            table_data.append(
                [
                    Paragraph(str(row.lab_sample_id), cell_style),
                    Paragraph(str(row.lab_no or ""), cell_style),
                    Paragraph(row.patient_name or "", cell_style),
                    Paragraph(row.patient_category or "", cell_style),
                    Paragraph(str(_format_value(row.taken_at) or ""), cell_style),
                    Paragraph(row.department_name or "", cell_style),
                    Paragraph(row.material_type or "", cell_style),
                    Paragraph(row.microorganism or "", cell_style),
                    Paragraph(row.antibiotic or "", cell_style),
                ]
            )

        file_path.parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=landscape(A4),
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        # Calculate optimal generic column widths for A4 landscape (277mm usable width roughly)
        available_width = landscape(A4)[0] - 20 * mm
        col_widths = [
            available_width * 0.05,  # ID
            available_width * 0.08,  # Lab No
            available_width * 0.15,  # Patient
            available_width * 0.08,  # Category
            available_width * 0.08,  # Date
            available_width * 0.12,  # Department
            available_width * 0.12,  # Material
            available_width * 0.15,  # Microorganism
            available_width * 0.17,  # Antibiotic
        ]

        summary_table = Table(summary_data, colWidths=[available_width * 0.3, available_width * 0.7])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTNAME", (0, 0), (-1, -1), unicode_font),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                ]
            )
        )
        filter_table = Table(filter_data, repeatRows=1, colWidths=[available_width * 0.3, available_width * 0.7])
        filter_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), unicode_font),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("TOPPADDING", (0, 0), (-1, -1), 1 * mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1 * mm),
                ]
            )
        )
        data_table = Table(table_data, repeatRows=1, colWidths=col_widths)
        data_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), unicode_font),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("TOPPADDING", (0, 0), (-1, -1), 1 * mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1 * mm),
                ]
            )
        )
        build_invariant_pdf(doc, [summary_table, filter_table, data_table])

        artifact_path = self._save_artifact_copy(report_type="analytics", source_path=file_path)
        report_hash = sha256_file(artifact_path)
        report_run_id = self._log_report_run(
            report_type="analytics",
            filters=filters,
            summary=agg,
            file_path=artifact_path,
            sha256=report_hash,
            created_by=actor_id,
        )
        return {
            "path": str(file_path),
            "artifact_path": str(artifact_path),
            "count": len(rows),
            "sha256": report_hash,
            "report_run_id": report_run_id,
        }

    def export_form100_pdf(
        self,
        *,
        card_id: str,
        file_path: str | Path,
        actor_id: int | None,
    ) -> dict[str, Any]:
        if self.form100_v2_service is None:
            raise ValueError("Form100 service is not configured")
        if actor_id is None:
            raise ValueError("actor_id обязателен для экспорта Form100 PDF")
        file_path = Path(file_path)
        render_result = self.form100_v2_service.export_pdf(card_id=card_id, file_path=file_path, actor_id=actor_id)
        report_type = "form100"

        artifact_path = self._save_artifact_copy(report_type=report_type, source_path=file_path)
        report_hash = sha256_file(artifact_path)
        report_run_id = self._log_report_run(
            report_type=report_type,
            filters={"card_id": card_id},
            summary={"path": str(render_result.get("path", file_path)), "card_id": card_id},
            file_path=artifact_path,
            sha256=report_hash,
            created_by=actor_id,
        )
        return {
            "path": str(file_path),
            "artifact_path": str(artifact_path),
            "sha256": report_hash,
            "report_run_id": report_run_id,
            "card_id": card_id,
        }

    def list_report_runs(
        self,
        *,
        limit: int = 100,
        report_type: str | None = None,
        query: str | None = None,
        verify_hash: bool = False,
    ) -> list[dict[str, Any]]:
        with self.session_factory() as session:
            stmt = select(models.ReportRun)
            if report_type:
                stmt = stmt.where(models.ReportRun.report_type == report_type)
            if query:
                like_pattern = f"%{query}%"
                stmt = stmt.where(
                    models.ReportRun.artifact_path.ilike(like_pattern)
                    | models.ReportRun.artifact_sha256.ilike(like_pattern)
                    | models.ReportRun.filters_json.ilike(like_pattern)
                )
            stmt = stmt.order_by(models.ReportRun.created_at.desc()).limit(limit)
            rows = list(session.execute(stmt).scalars())

        return [self._build_report_history_row(item, verify_hash=verify_hash) for item in rows]

    def verify_report_run(self, report_run_id: int) -> dict[str, Any]:
        with self.session_factory() as session:
            item = session.get(models.ReportRun, report_run_id)
            if item is None:
                raise ValueError("Запись отчета не найдена")
            artifact_path = cast(str | None, item.artifact_path)
            expected_sha256 = cast(str | None, item.artifact_sha256)
        return self._verify_artifact(
            report_run_id=report_run_id,
            artifact_path=artifact_path,
            expected_sha256=expected_sha256,
            compute_hash=True,
        )

    def _log_report_run(
        self,
        report_type: str,
        filters: dict,
        summary: dict,
        file_path: Path,
        sha256: str,
        created_by: int | None,
    ) -> int:
        safe_filters = self._sanitize_filters(cast(dict[str, Any], filters))
        with self.session_factory() as session:
            row = models.ReportRun(
                report_type=report_type,
                filters_json=self._json_dumps(safe_filters),
                result_summary_json=self._json_dumps(summary),
                artifact_path=str(file_path),
                artifact_sha256=sha256,
                created_by=created_by,
            )
            session.add(row)
            session.flush()
            return _as_int(row.id)

    def _sanitize_filters(self, filters: dict[str, Any]) -> dict[str, Any]:
        sanitized: dict[str, Any] = {}
        for key, value in filters.items():
            if key.lower() in SENSITIVE_FILTER_KEYS:
                sanitized[key] = "***"
                continue
            sanitized[key] = value
        return sanitized

    def _json_dumps(self, payload: dict[str, Any]) -> str:
        return json.dumps(payload, ensure_ascii=False, default=self._json_default)

    @staticmethod
    def _json_default(value: Any) -> str:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, Path):
            return str(value)
        return str(value)

    def _build_report_history_row(
        self,
        item: models.ReportRun,
        *,
        verify_hash: bool,
    ) -> dict[str, Any]:
        summary = self._safe_json_loads(cast(str, item.result_summary_json))
        filters = self._safe_json_loads(cast(str, item.filters_json))
        artifact_path = cast(str | None, item.artifact_path)
        artifact_sha256 = cast(str | None, item.artifact_sha256)
        verification = self._verify_artifact(
            report_run_id=_as_int(item.id),
            artifact_path=artifact_path,
            expected_sha256=artifact_sha256,
            compute_hash=verify_hash,
        )
        return {
            "id": _as_int(item.id),
            "created_at": item.created_at,
            "created_by": item.created_by,
            "report_type": item.report_type,
            "filters": filters,
            "summary": summary,
            "artifact_path": artifact_path,
            "artifact_sha256": artifact_sha256,
            "verification": verification,
        }

    def _verify_artifact(
        self,
        *,
        report_run_id: int,
        artifact_path: str | None,
        expected_sha256: str | None,
        compute_hash: bool,
    ) -> dict[str, Any]:
        result: dict[str, Any] = {
            "report_run_id": report_run_id,
            "artifact_path": artifact_path,
            "artifact_exists": False,
            "expected_sha256": expected_sha256,
            "actual_sha256": None,
            "verified": None,
            "status": "pending",
            "message": "Ожидает проверки",
        }
        if not artifact_path:
            result["status"] = "missing"
            result["verified"] = False
            result["message"] = "Артефакт не указан"
            return result

        path = Path(artifact_path)
        if not path.exists():
            result["status"] = "missing"
            result["verified"] = False
            result["message"] = "Файл артефакта не найден"
            return result

        result["artifact_exists"] = True
        if not compute_hash:
            result["status"] = "available"
            result["message"] = "Файл найден"
            return result

        if not expected_sha256:
            result["status"] = "error"
            result["verified"] = False
            result["message"] = "Эталонный SHA256 не сохранен"
            return result
        try:
            actual_sha256 = sha256_file(path)
        except OSError as exc:
            result["status"] = "error"
            result["verified"] = False
            result["message"] = f"Ошибка чтения файла: {exc}"
            return result

        result["actual_sha256"] = actual_sha256
        if actual_sha256 == expected_sha256:
            result["status"] = "ok"
            result["verified"] = True
            result["message"] = "SHA256 совпадает"
            return result

        result["status"] = "mismatch"
        result["verified"] = False
        result["message"] = "SHA256 не совпадает"
        return result

    def _save_artifact_copy(self, *, report_type: str, source_path: Path) -> Path:
        if not source_path.exists():
            raise FileNotFoundError(f"Файл отчета не найден: {source_path}")

        now = datetime.now(UTC)
        artifact_dir = REPORT_ARTIFACT_DIR / report_type / now.strftime("%Y") / now.strftime("%m")
        artifact_dir.mkdir(parents=True, exist_ok=True)

        suffix = source_path.suffix or ".bin"
        artifact_name = f"{report_type}_{now.strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}{suffix}"
        artifact_path = artifact_dir / artifact_name
        shutil.copy2(source_path, artifact_path)
        return artifact_path

    def _safe_json_loads(self, payload: str) -> dict[str, Any]:
        try:
            value = json.loads(payload)
        except (json.JSONDecodeError, TypeError):
            return {}
        if isinstance(value, dict):
            return value
        return {}

    def _build_filter_maps(self) -> dict[str, dict]:
        if not self.reference_service:
            return {}
        try:
            return {
                "department_id": {d.id: d.name for d in self.reference_service.list_departments()},
                "microorganism_id": {
                    m.id: f"{m.code or '-'} - {m.name}"
                    for m in self.reference_service.list_microorganisms()
                },
                "antibiotic_id": {
                    a.id: f"{a.code} - {a.name}" for a in self.reference_service.list_antibiotics()
                },
                "material_type_id": {
                    m.id: f"{m.code} - {m.name}"
                    for m in self.reference_service.list_material_types()
                },
                "icd10_code": {
                    i.code: f"{i.code} - {i.title}" for i in self.reference_service.list_icd10()
                },
            }
        except (AttributeError, SQLAlchemyError, TypeError, ValueError) as exc:
            logger.warning("Failed to build filter maps for report export: %s", exc)
            return {}

    def _format_filter_value(self, key: str, value: Any, filter_maps: dict[str, dict]) -> str:
        if key == "growth_flag":
            if value == 1:
                return "Да"
            if value == 0:
                return "Нет"
        if key == "patient_category":
            if isinstance(value, MilitaryCategory):
                return value.value
            if isinstance(value, str):
                try:
                    return MilitaryCategory[value].value
                except KeyError:
                    return value
        if key in filter_maps:
            mapped = filter_maps[key].get(value)
            if mapped is not None:
                return str(mapped)
        return str(_format_value(value))


