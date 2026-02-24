from __future__ import annotations

import json
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def load_form100_excel(file_path: Path) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    cards = _read_sheet(wb, "form100_card")
    marks = _read_sheet(wb, "form100_mark")
    stages = _read_sheet(wb, "form100_stage")
    return {"cards": cards, "marks": marks, "stages": stages}


def _read_sheet(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    row_iter = ws.iter_rows(values_only=True)
    header = next(row_iter, None)
    if not header:
        return []
    header_names = [str(value) if value is not None else "" for value in header]
    rows: list[dict[str, Any]] = []
    for row in row_iter:
        payload = {
            key: _parse_cell_value(key, row[idx] if idx < len(row) else None)
            for idx, key in enumerate(header_names)
            if key
        }
        rows.append(payload)
    return rows


def _parse_cell_value(key: str, value: Any) -> Any:
    if value is None or value == "":
        return None
    if key in {"trauma_types", "wound_types", "features", "shape_json", "meta_json"}:
        if isinstance(value, str):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                return [] if key in {"trauma_types", "wound_types", "features"} else {}
        return value
    if key in {
        "birth_date",
        "outcome_date",
    }:
        return _parse_date(value)
    if key in {
        "created_at",
        "updated_at",
        "injury_dt",
        "arrival_dt",
        "signed_at",
        "received_at",
        "evac_next_dt",
    }:
        return _parse_datetime(value)
    if key in {
        "first_aid_before",
        "is_combat",
        "flag_urgent",
        "flag_sanitation",
        "flag_isolation",
        "flag_radiation",
        "care_analgesia_given",
        "care_antibiotic_given",
        "care_antidote_given",
        "infusion_performed",
        "transfusion_performed",
        "sanitation_performed",
        "evac_require_escort",
        "evac_oxygen_needed",
        "seal_applied",
    }:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(int(value))
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "yes", "да"}
    return value


def _parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        text = value.strip()
        if text.endswith("Z"):
            text = f"{text[:-1]}+00:00"
        try:
            parsed = datetime.fromisoformat(text)
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=UTC)
            return parsed
        except ValueError:
            pass
        date_and_time = text.split(" ")
        if len(date_and_time) == 2 and "." in date_and_time[0]:
            date_parts = date_and_time[0].split(".")
            time_parts = date_and_time[1].split(":")
            if len(date_parts) == 3 and len(time_parts) >= 2:
                try:
                    day, month, year = (int(part) for part in date_parts)
                    hour = int(time_parts[0])
                    minute = int(time_parts[1])
                    second = int(time_parts[2]) if len(time_parts) > 2 else 0
                    return datetime(year, month, day, hour, minute, second, tzinfo=UTC)
                except ValueError:
                    return None
    return None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            pass
        parts = value.split(".")
        if len(parts) == 3:
            try:
                day, month, year = (int(part) for part in parts)
                return date(year, month, day)
            except ValueError:
                return None
    return None
