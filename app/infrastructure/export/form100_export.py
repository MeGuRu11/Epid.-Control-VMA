from __future__ import annotations

import json
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.infrastructure.security.sha256 import sha256_file

FORM100_EXPORT_SCHEMA = "form100.package.v1"

CARD_SHEET = "form100_card"
MARK_SHEET = "form100_mark"
STAGE_SHEET = "form100_stage"

CARD_COLUMNS: list[str] = [
    "id",
    "created_at",
    "updated_at",
    "created_by",
    "updated_by",
    "status",
    "version",
    "qr_payload",
    "print_number",
    "corrects_id",
    "corrected_by_new_id",
    "last_name",
    "first_name",
    "middle_name",
    "birth_date",
    "rank",
    "unit",
    "dog_tag_number",
    "id_doc_type",
    "id_doc_number",
    "injury_dt",
    "arrival_dt",
    "first_aid_before",
    "cause_category",
    "is_combat",
    "trauma_types",
    "thermal_degree",
    "wound_types",
    "features",
    "other_text",
    "diagnosis_text",
    "diagnosis_code",
    "triage",
    "flag_urgent",
    "flag_sanitation",
    "flag_isolation",
    "flag_radiation",
    "care_bleeding_control",
    "care_dressing",
    "care_immobilization",
    "care_airway",
    "care_analgesia_given",
    "care_analgesia_details",
    "care_antibiotic_given",
    "care_antibiotic_details",
    "care_antidote_given",
    "care_antidote_details",
    "care_tetanus",
    "care_other",
    "infusion_performed",
    "infusion_volume_ml",
    "infusion_details",
    "transfusion_performed",
    "transfusion_volume_ml",
    "transfusion_details",
    "sanitation_performed",
    "sanitation_type",
    "sanitation_details",
    "evac_destination",
    "evac_transport",
    "evac_position",
    "evac_require_escort",
    "evac_oxygen_needed",
    "evac_notes",
    "signed_by",
    "signed_at",
    "seal_applied",
]

MARK_COLUMNS: list[str] = [
    "id",
    "card_id",
    "side",
    "type",
    "shape_json",
    "meta_json",
    "created_at",
    "created_by",
]

STAGE_COLUMNS: list[str] = [
    "id",
    "card_id",
    "stage_name",
    "received_at",
    "updated_diagnosis_text",
    "updated_diagnosis_code",
    "procedures_text",
    "evac_next_destination",
    "evac_next_dt",
    "condition_at_transfer",
    "outcome",
    "outcome_date",
    "burial_place",
    "signed_by",
    "signed_at",
]


def export_form100_excel(
    *,
    cards: list[dict[str, Any]],
    marks: list[dict[str, Any]],
    stages: list[dict[str, Any]],
    file_path: Path,
) -> dict[str, int]:
    wb = Workbook()

    meta = wb.active
    meta.title = "meta"
    meta.append(["schema", FORM100_EXPORT_SCHEMA])
    meta.append(["exported_at", datetime.now(UTC).isoformat()])

    card_ws = wb.create_sheet(title=CARD_SHEET)
    card_ws.append(CARD_COLUMNS)
    for row in cards:
        card_ws.append([_serialize_value(row.get(col)) for col in CARD_COLUMNS])

    mark_ws = wb.create_sheet(title=MARK_SHEET)
    mark_ws.append(MARK_COLUMNS)
    for row in marks:
        mark_ws.append([_serialize_value(row.get(col)) for col in MARK_COLUMNS])

    stage_ws = wb.create_sheet(title=STAGE_SHEET)
    stage_ws.append(STAGE_COLUMNS)
    for row in stages:
        stage_ws.append([_serialize_value(row.get(col)) for col in STAGE_COLUMNS])

    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    return {
        CARD_SHEET: len(cards),
        MARK_SHEET: len(marks),
        STAGE_SHEET: len(stages),
    }


def build_manifest(*, files: list[Path], exported_by: str | None = None) -> dict[str, Any]:
    entries = []
    for file_path in files:
        entries.append(
            {
                "name": file_path.name,
                "sha256": sha256_file(file_path),
                "size": file_path.stat().st_size,
            }
        )
    return {
        "schema_version": FORM100_EXPORT_SCHEMA,
        "exported_at": datetime.now(UTC).isoformat(),
        "exported_by": exported_by,
        "files": entries,
    }


def _serialize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value
