from __future__ import annotations

import csv
from collections.abc import Mapping
from pathlib import Path
from typing import cast

from openpyxl import load_workbook
from PySide6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QWizard,
    QWizardPage,
)

from app.application.dto.auth_dto import SessionContext
from app.application.security import can_manage_exchange
from app.application.services.exchange_service import ExchangeService
from app.ui.widgets.async_task import run_async
from app.ui.widgets.button_utils import compact_button
from app.ui.widgets.dialog_utils import exec_message_box
from app.ui.widgets.notifications import show_error, show_info, show_warning
from app.ui.widgets.table_utils import connect_combo_autowidth, resize_columns_to_content


class ImportExportWizard(QWizard):
    def __init__(
        self,
        exchange_service: ExchangeService,
        session: SessionContext,
        table_labels: dict[str, str],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.exchange_service = exchange_service
        self.session = session
        self.table_labels = table_labels

        self.setWindowTitle("Мастер импорта/экспорта")
        self.setButtonText(QWizard.WizardButton.BackButton, "Назад")
        self.setButtonText(QWizard.WizardButton.NextButton, "Далее")
        self.setButtonText(QWizard.WizardButton.FinishButton, "Готово")
        self.setButtonText(QWizard.WizardButton.CancelButton, "Отмена")
        self._direction_page = DirectionPage(self)
        self._path_page = PathPage(self)
        self._preview_page = PreviewPage(self)
        self.addPage(self._direction_page)
        self.addPage(self._path_page)
        self.addPage(self._preview_page)

    @property
    def operation_host(self) -> QWidget:
        parent = self.parentWidget()
        return parent if parent is not None else self

    def _ensure_permissions(self) -> None:
        if not can_manage_exchange(self.session.role):
            raise ValueError("Недостаточно прав для операций импорта/экспорта")

    def accept(self) -> None:
        self._ensure_permissions()
        direction = self._direction_page.direction.currentData()
        fmt = self._direction_page.format.currentData()
        table_name = self._direction_page.table_select.currentData()
        import_mode = self._direction_page.import_mode.currentData()
        file_path = self._path_page.path_input.text().strip()

        if not file_path:
            show_warning(self, "Укажите путь к файлу")
            return
        if direction == "export":
            reply = exec_message_box(
                self,
                "Внимание: персональные данные",
                (
                    "Экспортируемый файл может содержать персональные и медицинские данные.\n"
                    "Обеспечьте безопасное хранение и передачу файла.\n\n"
                    "Продолжить экспорт?"
                ),
                icon=QMessageBox.Icon.Warning,
                buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                default_button=QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        self._set_busy(True)

        def _run() -> tuple[str, bool]:
            return self._run_operation(direction, fmt, table_name, file_path, import_mode)

        def _on_success(result: tuple[str, bool]) -> None:
            message, has_errors = result
            if direction == "export":
                show_info(self.operation_host, f"Экспорт завершён: {message}")
            else:
                if has_errors:
                    show_warning(self.operation_host, f"Импорт завершён с ошибками: {message}")
                else:
                    show_info(self.operation_host, f"Импорт завершён: {message}")
            self._accept_success()

        def _on_error(exc: Exception) -> None:
            show_error(self, str(exc))

        run_async(
            self,
            _run,
            on_success=_on_success,
            on_error=_on_error,
            on_finished=lambda: self._set_busy(False),
        )

    def _set_busy(self, busy: bool) -> None:
        for button in (
            QWizard.WizardButton.BackButton,
            QWizard.WizardButton.NextButton,
            QWizard.WizardButton.FinishButton,
            QWizard.WizardButton.CancelButton,
        ):
            btn = self.button(button)
            if btn:
                btn.setEnabled(not busy)

    def _accept_success(self) -> None:
        super().accept()

    def _run_operation(
        self,
        direction: str,
        fmt: str,
        table_name: str | None,
        file_path: str,
        import_mode: str,
    ) -> tuple[str, bool]:
        self._ensure_permissions()
        actor_id = self.session.user_id
        if direction == "export":
            if fmt == "excel":
                excel_result = self.exchange_service.export_excel(
                    file_path=file_path,
                    exported_by=self.session.login,
                    actor_id=actor_id,
                )
                total = sum(excel_result["counts"].values())
                return f"{total} записей", False
            if fmt == "csv":
                if not table_name:
                    raise ValueError("Выберите таблицу для CSV")
                csv_result = self.exchange_service.export_csv(
                    file_path=file_path,
                    table_name=table_name,
                    actor_id=actor_id,
                )
                return f"{csv_result['count']} записей", False
            if fmt == "pdf":
                if not table_name:
                    raise ValueError("Выберите таблицу для PDF")
                pdf_result = self.exchange_service.export_pdf(
                    file_path=file_path,
                    table_name=table_name,
                    actor_id=actor_id,
                )
                return f"{pdf_result['count']} записей", False
            if fmt == "zip":
                zip_result = self.exchange_service.export_zip(
                    file_path=file_path,
                    exported_by=self.session.login,
                    actor_id=actor_id,
                )
                total = sum(zip_result["counts"].values())
                return f"{total} записей", False
            if fmt == "form100_zip":
                form100_result = self.exchange_service.export_form100_package_zip(
                    file_path=file_path,
                    exported_by=self.session.login,
                    actor_id=actor_id,
                    card_id=None,
                )
                counts = cast(dict[str, int], cast(dict[str, object], form100_result).get("counts", {}))
                total = sum(counts.values())
                return f"{total} записей", False
        else:
            if fmt == "excel":
                excel_import_result = self.exchange_service.import_excel(
                    file_path=file_path, actor_id=actor_id, mode=import_mode
                )
                return self._format_import_result(excel_import_result)
            if fmt == "csv":
                if not table_name:
                    raise ValueError("Выберите таблицу для CSV")
                csv_import_result = self.exchange_service.import_csv(
                    file_path=file_path,
                    table_name=table_name,
                    actor_id=actor_id,
                    mode=import_mode,
                )
                return self._format_import_result(csv_import_result)
            if fmt == "zip":
                zip_import_result = self.exchange_service.import_zip(
                    file_path=file_path,
                    actor_id=actor_id,
                    mode=import_mode,
                )
                return self._format_import_result(zip_import_result)
            if fmt == "form100_zip":
                form100_import_result = self.exchange_service.import_form100_package_zip(
                    file_path=file_path,
                    actor_id=actor_id,
                    mode=import_mode,
                )
                return self._format_import_result(cast(Mapping[str, object], form100_import_result))

        raise ValueError("Неподдерживаемый формат операции")

    def _format_import_result(self, result: Mapping[str, object]) -> tuple[str, bool]:
        summary = cast(dict[str, object], result.get("summary") or {})
        rows_total = _to_int(summary.get("rows_total"))
        added = _to_int(summary.get("added"))
        updated = _to_int(summary.get("updated"))
        skipped = _to_int(summary.get("skipped"))
        errors = _to_int(summary.get("errors"))
        details = cast(dict[str, dict[str, object]], result.get("details") or {})
        details_items = []
        for scope, values in details.items():
            rows = _to_int(values.get("rows"))
            scope_added = _to_int(values.get("added"))
            scope_updated = _to_int(values.get("updated"))
            scope_skipped = _to_int(values.get("skipped"))
            scope_errors = _to_int(values.get("errors"))
            details_items.append(
                f"{scope}: rows={rows}, +{scope_added}/~{scope_updated}/-{scope_skipped}/!{scope_errors}"
            )
        message_parts = [
            (
                f"rows={rows_total}, добавлено={added}, обновлено={updated}, "
                f"пропущено={skipped}, ошибок={errors}"
            )
        ]
        if details_items:
            message_parts.append("детали: " + "; ".join(details_items))
        error_log_path = result.get("error_log_path")
        if error_log_path:
            message_parts.append(f"лог ошибок: {error_log_path}")
        return " | ".join(message_parts), errors > 0


def _to_int(value: object) -> int:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        try:
            return int(value)
        except ValueError:
            return 0
    return 0


class DirectionPage(QWizardPage):
    def __init__(self, wizard: ImportExportWizard) -> None:
        super().__init__(wizard)
        self.wizard_ref = wizard
        self.setTitle("Выбор операции")
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QFormLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setVerticalSpacing(10)
        self.direction = QComboBox()
        self.direction.addItem("Экспорт", "export")
        self.direction.addItem("Импорт", "import")
        connect_combo_autowidth(self.direction)
        self.direction.currentIndexChanged.connect(self._sync_state)

        self.format = QComboBox()
        self.format.addItem("Excel", "excel")
        self.format.addItem("CSV", "csv")
        self.format.addItem("PDF", "pdf")
        self.format.addItem("ZIP", "zip")
        self.format.addItem("Form100 ZIP", "form100_zip")
        connect_combo_autowidth(self.format)
        self.format.currentIndexChanged.connect(self._sync_state)

        self.table_select = QComboBox()
        self.table_select.addItem("Выбрать", None)
        for key, label in self.wizard_ref.table_labels.items():
            self.table_select.addItem(label, key)
        connect_combo_autowidth(self.table_select)

        self.import_mode = QComboBox()
        self.import_mode.addItem("Обновление/слияние", "merge")
        self.import_mode.addItem("Только добавлять", "append")
        connect_combo_autowidth(self.import_mode)

        layout.addRow("Операция", self.direction)
        layout.addRow("Формат", self.format)
        layout.addRow("Таблица (для CSV/PDF)", self.table_select)
        layout.addRow("Режим импорта", self.import_mode)
        self._table_label = layout.labelForField(self.table_select)
        self._import_mode_label = layout.labelForField(self.import_mode)
        self._context_hint = QLabel("")
        self._context_hint.setWordWrap(True)
        self._context_hint.setObjectName("muted")
        layout.addRow("", self._context_hint)
        self._sync_state()

    def _sync_state(self) -> None:
        is_import = self.direction.currentData() == "import"
        fmt = self.format.currentData()
        if is_import and fmt == "pdf":
            self.format.setCurrentIndex(0)
            fmt = self.format.currentData()
        table_visible = fmt in {"csv", "pdf"}
        import_mode_visible = is_import
        if self._table_label is not None:
            self._table_label.setVisible(table_visible)
        self.table_select.setVisible(table_visible)
        self.table_select.setEnabled(table_visible)
        if self._import_mode_label is not None:
            self._import_mode_label.setVisible(import_mode_visible)
        self.import_mode.setVisible(import_mode_visible)
        self.import_mode.setEnabled(import_mode_visible)
        self._context_hint.setText(self._build_context_hint(is_import, str(fmt)))

    def _build_context_hint(self, is_import: bool, fmt: str) -> str:
        if fmt == "csv":
            if is_import:
                return "CSV импортируется по одной таблице. Выберите таблицу и режим загрузки."
            return "CSV экспортируется по одной таблице. Выберите раздел, который нужно выгрузить."
        if fmt == "pdf":
            return "PDF выгружается по одной таблице. Выберите раздел для печатного отчёта."
        if fmt == "excel":
            if is_import:
                return "Excel импортирует весь пакет данных. Режим определяет, обновлять ли существующие записи."
            return "Excel экспортирует полный набор листов с русскими названиями и понятными заголовками."
        if fmt == "zip":
            if is_import:
                return "ZIP импортирует полный пакет обмена. Можно обновлять записи или только добавлять новые."
            return "ZIP экспортирует полный пакет обмена со служебным manifest и Excel-файлом."
        if fmt == "form100_zip":
            return "Form100 ZIP работает как архив специализированного обмена карточками Формы 100."
        return ""


class PathPage(QWizardPage):
    def __init__(self, wizard: ImportExportWizard) -> None:
        super().__init__(wizard)
        self.wizard_ref = wizard
        self.setTitle("Файл")
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QHBoxLayout(self)
        self.path_input = QLineEdit()
        browse_btn = QPushButton("Выбрать файл")
        compact_button(browse_btn)
        browse_btn.clicked.connect(self._browse)
        layout.addWidget(self.path_input)
        layout.addWidget(browse_btn)

    def _browse(self) -> None:
        direction = self.wizard_ref._direction_page.direction.currentData()
        fmt = self.wizard_ref._direction_page.format.currentData()

        if direction == "export":
            filename = "export"
            if fmt == "excel":
                path, _ = QFileDialog.getSaveFileName(self, "Экспорт Excel", f"{filename}.xlsx", "Excel (*.xlsx)")
            elif fmt == "csv":
                path, _ = QFileDialog.getSaveFileName(self, "Экспорт CSV", f"{filename}.csv", "CSV (*.csv)")
            elif fmt == "pdf":
                path, _ = QFileDialog.getSaveFileName(self, "Экспорт PDF", f"{filename}.pdf", "PDF (*.pdf)")
            elif fmt == "form100_zip":
                path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Экспорт Form100 ZIP",
                    "form100_export.zip",
                    "ZIP (*.zip)",
                )
            else:
                path, _ = QFileDialog.getSaveFileName(self, "Экспорт ZIP", f"{filename}.zip", "ZIP (*.zip)")
        else:
            if fmt == "excel":
                path, _ = QFileDialog.getOpenFileName(self, "Импорт Excel", "", "Excel (*.xlsx)")
            elif fmt == "csv":
                path, _ = QFileDialog.getOpenFileName(self, "Импорт CSV", "", "CSV (*.csv)")
            elif fmt == "form100_zip":
                path, _ = QFileDialog.getOpenFileName(self, "Импорт Form100 ZIP", "", "ZIP (*.zip)")
            else:
                path, _ = QFileDialog.getOpenFileName(self, "Импорт ZIP", "", "ZIP (*.zip)")
        if path:
            self.path_input.setText(path)


class PreviewPage(QWizardPage):
    def __init__(self, wizard: ImportExportWizard) -> None:
        super().__init__(wizard)
        self.wizard_ref = wizard
        self.setTitle("Предпросмотр и проверка")
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        self.summary_label = QLabel("Проверьте параметры и файл перед запуском.")
        self.summary_label.setWordWrap(True)
        layout.addWidget(self.summary_label)
        self.preview_table = QTableWidget(0, 0)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.verticalHeader().setVisible(False)
        layout.addWidget(self.preview_table)

    def initializePage(self) -> None:  # noqa: N802
        direction = self.wizard_ref._direction_page.direction.currentData()
        fmt = self.wizard_ref._direction_page.format.currentData()
        table_name = self.wizard_ref._direction_page.table_select.currentData()
        table_text = self.wizard_ref._direction_page.table_select.currentText()
        mode_text = self.wizard_ref._direction_page.import_mode.currentText()
        file_path = self.wizard_ref._path_page.path_input.text().strip()
        direction_text = self.wizard_ref._direction_page.direction.currentText()
        format_text = self.wizard_ref._direction_page.format.currentText()
        summary = f"Операция: {direction_text}, формат: {format_text}"
        if table_name:
            summary += f", таблица: {table_text}"
        if direction == "import":
            summary += f", режим: {mode_text}"
        if file_path:
            summary += f", файл: {file_path}"
        self.summary_label.setText(summary)

        self.preview_table.clearContents()
        self.preview_table.setRowCount(0)
        self.preview_table.setColumnCount(0)

        if direction == "export":
            return
        if not file_path:
            return
        path = Path(file_path)
        if not path.exists():
            return
        if fmt == "csv":
            self._preview_csv(path)
        elif fmt == "excel":
            self._preview_excel(path)

    def _preview_csv(self, path: Path) -> None:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            rows = []
            for _ in range(20):
                try:
                    rows.append(next(reader))
                except StopIteration:
                    break
        if not rows:
            return
        self.preview_table.setColumnCount(len(rows[0]))
        self.preview_table.setHorizontalHeaderLabels(rows[0])
        for row_idx, row in enumerate(rows[1:]):
            self.preview_table.insertRow(row_idx)
            for col_idx, value in enumerate(row):
                self.preview_table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))
        resize_columns_to_content(self.preview_table)

    def _preview_excel(self, path: Path) -> None:
        wb = load_workbook(path, read_only=True)
        ws = next((wb[name] for name in wb.sheetnames if name != "meta"), wb.active)
        if ws is None:
            return
        rows = list(ws.iter_rows(values_only=True))[:20]
        if not rows:
            return
        header = [str(x) if x is not None else "" for x in rows[0]]
        self.preview_table.setColumnCount(len(header))
        self.preview_table.setHorizontalHeaderLabels(header)
        for row_idx, row in enumerate(rows[1:]):
            self.preview_table.insertRow(row_idx)
            for col_idx, value in enumerate(row):
                self.preview_table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))
        resize_columns_to_content(self.preview_table)

